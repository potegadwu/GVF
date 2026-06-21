import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Proces Maj-Czerwiec z podziałem"

# Style
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
month_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
month_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
total_font = Font(name="Calibri", bold=True, size=11, color="006100")
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Nagłówki
headers = ["Kategoria", "Działanie", "Status", "Termin", "Godziny (h)", "Uwagi"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

col_widths = [35, 45, 16, 22, 12, 35]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

def add_month_header(ws, row_num, month_name):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    cell = ws.cell(row=row_num, column=1, value=month_name)
    cell.font = month_font
    cell.fill = month_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 7):
        ws.cell(row=row_num, column=col).fill = month_fill
        ws.cell(row=row_num, column=col).border = thin_border
    return row_num + 1

def add_row(ws, row_num, data, hours=None):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 5 and hours:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    return row_num + 1

def add_total(ws, row_num, label, total_hours):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal='right', vertical='center')
    for col in range(1, 7):
        ws.cell(row=row_num, column=col).border = thin_border
        ws.cell(row=row_num, column=col).fill = total_fill
    total_cell = ws.cell(row=row_num, column=5, value=total_hours)
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.alignment = Alignment(horizontal='center', vertical='center')
    return row_num + 1

row = 2

# ============================================================
# CZERWIEC 2026
# ============================================================
row = add_month_header(ws, row, "CZERWIEC 2026")

# Zadania z terminami w czerwcu
czerwiec = [
    ["Urodziny LU", "Prezentacja konceptu urodzinowego", "Zrealizowano", "Do 23.06", 8, "Omawiane 12.06, prezentacja 16.06"],
    ["Urodziny LU", "Wycena wydarzenia (drony, lasery, gadżety)", "W trakcie", "Koniec czerwca", 4, "Zapytania wysłane, odpowiedzi z końcem tygodnia"],
    ["Urodziny LU", "Weryfikacja terminu 7 września", "Do realizacji", "Do 23.06", None, "TBC"],
    ["Urodziny LU", "Zaproszenie / Save the date", "Do realizacji", "Wysyłka ~26.06", 8, "Pierwsze wrażenie - nie spieszyć się"],
    ["Urodziny LU", "Komunikacja eventu (maile)", "Zaplanowano", "20 czerwca 1st mailing", 10, "Lipiec i sierpień: przypominające"],
    ["Urodziny LU", "Menu na wydarzenie", "Do realizacji", "Do 23.06", 4, "Marcin W ma przygotować propozycję"],
    ["Urodziny LU", "Scenariusz / Prowadzący", "Do realizacji", "Do końca czerwca", 8, "Decyzja kto prowadzi, wycena prowadzącego"],
    ["Urodziny LU", "Rozpoznanie partnerstwa (właściciel budynku)", "W trakcie", "Do końca czerwca", 1, "Paweł ma rozeznać temat"],
    ["Urodziny LU", "Budżet wydarzenia", "Do realizacji", "Do końca czerwca", 6, "Policzenie kosztów"],
    ["Logo grupy", "Logo grupy - projekt finalny", "W trakcie", "Do 17.06", 12, "Wybraliśmy kierunek, we wtorek finalne"],
    ["Wizytówki", "Wizytówki - wycena i druk", "W trakcie", "Do końca czerwca", 4, "Wybrane kartony - dłuższy czas oczekiwania"],
    ["Nowa formuła video", "Zmiana formuły materiałów video", "W trakcie", "Czerwiec", 10, "Spotkanie Michała z Marcinem 17.06"],
    ["Nowa formuła video", "Scenariusze postów na konta", "W trakcie", "Czerwiec", 3, "Posty idą na bieżąco"],
    ["Nowa formuła video", "Posty z sesji zeszłorocznej w willi", "W trakcie", "Czerwiec", 6, "Materiał 'roboczy', zabawny"],
    ["Kampanie reklamowe", "Media plan 2026 - rewizja", "W trakcie", "Czerwiec", 4, "100k zł netto, korekty w ramach budżetu"],
    ["Kampanie reklamowe", "Reklamy z Robertem, Bartkiem, Łukaszem", "W trakcie", "Czerwiec", 50, "Przygotowanie 30 kreacji do kampanii Meta"],
    ["Sesja VF", "Sesja zdjęciowa - ogród i kuchnia", "Zaplanowano", "24 czerwca", 5, "Scenariusz od Marcina C"],
    ["Nagrywki", "Wina gruzińskie (11 czerwca)", "Zrealizowano", "11.06", None, "Obszerna relacja, wszyscy managerowie"],
    ["Nagrywki", "Konkurs gotowania dziennikarzy (15.06)", "Zrealizowano", "15.06", None, "Poszło w Pytanie na Śniadanie"],
    ["Nagrywki", "Koncert Monika Borzym (19.06)", "Zaplanowano", "19.06", None, "Serwowane menu"],
    ["Nagrywki", "Gala kosmetyków (25.06)", "Zaplanowano", "25.06", None, "Start 19:00"],
    ["Nagrywki", "LU - Hebe + kosmetyki koreańskie", "Do potwierdzenia", "23.06", None, "Do potwierdzenia"],
]

czerwiec_total = 0
for item in czerwiec:
    row = add_row(ws, row, item, item[4])
    if item[4]:
        czerwiec_total += item[4]

row = add_total(ws, row, "RAZEM CZERWIEC 2026:", czerwiec_total)
row += 1

# ============================================================
# MAJ 2026
# ============================================================
row = add_month_header(ws, row, "MAJ 2026")

maj = [
    ["Pracodawcy RP", "Śniadanie u Pracodawcy RP", "Zrealizowano", "18-19.05", 2, "Paweł mówi o poszczególnych podmiotach"],
    ["Pracodawcy RP", "Mailing w Pracodawcy RP", "W trakcie", "Maj", 2, "Będziemy w mailingu"],
    ["Pracodawcy RP", "Materiał w miesięczniku", "W trakcie", "Maj", 2, "Przygotowanie treści"],
    ["Nagrywki", "Wesele w LU (30 maja)", "Zrealizowano", "30.05", None, "Nagranie materiału"],
    ["Nagrywki", "Manager - logistyka (28 maja)", "Zrealizowano", "28.05", None, "Dwa cateringi + impreza VF"],
    ["Identyfikacja grupy", "Analiza rynku, konkurencji, strategia", "W trakcie", "Maj", 20, "Omawiane 16.05"],
    ["Identyfikacja grupy", "Zmiana firmy na Villa Foksal Group", "W trakcie", "Maj", 20, "Omawiane 11.05"],
    ["Identyfikacja grupy", "Narracja grupy - 'Gościnność...'", "W trakcie", "Maj", 10, "Bazowa treść gotowa"],
    ["Identyfikacja grupy", "Sąsiad - Pałac na Foksal", "Zrealizowano", "Maj", 1, "Brak zagrożenia"],
]

maj_total = 0
for item in maj:
    row = add_row(ws, row, item, item[4])
    if item[4]:
        maj_total += item[4]

row = add_total(ws, row, "RAZEM MAJ 2026:", maj_total)
row += 1

# ============================================================
# LIPiec-SIERPIEŃ (wykraczające poza maj-czerwiec)
# ============================================================
row = add_month_header(ws, row, "LIPiec-SIERPIEŃ 2026 (wykracza poza zakres)")

poza_zakresem = [
    ["Urodziny LU", "Animacje na ekranach (diody, totemy, ekrany)", "Do realizacji", "Lipiec-sierpień", 100, "Wykorzystamy diode, totem i ekrany"],
    ["Urodziny LU", "Welcome / Good Night Pack", "Do realizacji", "Lipiec-sierpień", 2, "Coś użytecznego z logo grupy"],
    ["Urodziny LU", "Partnerstwo / Sponsorzy", "Do przemyślenia", "Lipiec", 6, "Partner niekonkurujący"],
    ["Urodziny LU", "Komunikacja eventu (maile)", "Zaplanowano", "Lipiec-sierpień", 10, "Mailingi przypominające"],
    ["CRM", "Definicja wymagań i specyfikacja", "Do realizacji", "Przed 6 lipca", 5, "Stworzyć dokument specyfikacji"],
    ["CRM", "Warsztaty CRM (2 spotkania)", "Zaplanowano", "6 i 13 lipca", 28, "Przygotowanie warsztatu + 2 spotkania po 5h"],
    ["Warsztaty strategiczne", "Warsztat 1", "Zaplanowano", "29 czerwca / 6 lipca", 5, "Strategia komunikacji, oferta, pozycjonowanie"],
    ["Warsztaty strategiczne", "Warsztat 2", "Zaplanowano", "6 lipca / 13 lipca", 5, "Kontynuacja prac"],
    ["Warsztaty strategiczne", "Spotkanie 'Kim jesteśmy na rynku'", "Zaplanowano", "Lipiec-sierpień", 6, "Komunikacja wewnętrzna"],
    ["Strona www", "Strona parasolowa VillaFoksalGroup.pl", "W trakcie", "Czerwiec-Lipiec", 30, "Projekt statyczny + podstrony"],
    ["Strona www", "Statyczne projekty stron LU i VF", "W trakcie", "Czerwiec", None, "Do zderzenia z zespołem"],
    ["Logo grupy", "Logo na materiałach i www", "Zaplanowano", "Po akceptacji", 3, "Pokazanie jak logo pracuje"],
    ["Forbes", "Współpraca z Forbes", "Oczekiwanie", None, 1, "Czekamy na ofertę"],
    ["Domeny", "Zakup domen dodatkowych", "W trakcie", None, 2, "VillaFoksalGroup.pl itp."],
]

poza_total = 0
for item in poza_zakresem:
    row = add_row(ws, row, item, item[4])
    if item[4]:
        poza_total += item[4]

row = add_total(ws, row, "RAZEM LIPiec-SIERPIEŃ:", poza_total)
row += 1

# ============================================================
# PODSUMOWANIE
# ============================================================
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
podsumowanie_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
podsumowanie_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
cell = ws.cell(row=row, column=1, value="PODSUMOWANIE - OKRES MAJ-CZERWIEC 2026")
cell.font = podsumowanie_font
cell.fill = podsumowanie_fill
cell.alignment = Alignment(horizontal='center', vertical='center')
for col in range(1, 7):
    ws.cell(row=row, column=col).fill = podsumowanie_fill
    ws.cell(row=row, column=col).border = thin_border

row += 1
podsumowanie = [
    ["CZERWIEC 2026", czerwiec_total],
    ["MAJ 2026", maj_total],
    ["LIPiec-SIERPIEŃ (poza zakresem)", poza_total],
]
for item in podsumowanie:
    cell_cat = ws.cell(row=row, column=1, value=item[0])
    cell_cat.font = bold_font
    cell_cat.border = thin_border
    cell_h = ws.cell(row=row, column=5, value=item[1])
    cell_h.font = bold_font
    cell_h.alignment = Alignment(horizontal='center', vertical='center')
    cell_h.border = thin_border
    for col in range(2, 7):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
razem_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
razem_font = Font(name="Calibri", bold=True, size=14, color="006100")
cell = ws.cell(row=row, column=1, value="RAZEM MAJ + CZERWIEC:")
cell.font = razem_font
cell.fill = razem_fill
cell.border = thin_border
cell_h = ws.cell(row=row, column=5, value=maj_total + czerwiec_total)
cell_h.font = razem_font
cell_h.fill = razem_fill
cell_h.alignment = Alignment(horizontal='center', vertical='center')
cell_h.border = thin_border
for col in range(1, 7):
    ws.cell(row=row, column=col).fill = razem_fill
    ws.cell(row=row, column=col).border = thin_border

ws.freeze_panes = 'A2'

# Zapis
output_path = r"C:\Users\Malgorzata\Desktop\GRUPA VILLA FOKSAL ANTYGRAVITY\GVF\GVF_Proces_Maj_Czerwiec_z_podzialem.xlsx"
wb.save(output_path)

print(f"Plik zapisany: {output_path}")
print(f"\nPODSUMOWANIE:")
print(f"  CZERWIEC 2026:        {czerwiec_total}h")
print(f"  MAJ 2026:             {maj_total}h")
print(f"  LIPiec-SIERPIEŃ:      {poza_total}h (poza zakresem)")
print(f"  ----------------------------------------")
print(f"  RAZEM MAJ+CZERWIEC:   {maj_total + czerwiec_total}h")
print(f"  RAZEM ZE WSZYSTKIM:   {maj_total + czerwiec_total + poza_total}h")
