import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: ZESTAWIENIE DZIAŁAŃ
# ============================================================
ws = wb.active
ws.title = "Zestawienie Działalń"

# Kolory
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
category_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
category_font = Font(name="Calibri", bold=True, size=11)
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
status_realized = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
status_progress = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
status_planned = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
status_blocked = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Nagłówki
headers = [
    "L.p.",
    "Kategoria",
    "Temat / Działanie",
    "Status",
    "Data rozpoczęcia",
    "Termin / Data realizacji",
    "Czas trwania",
    "Osoba odpowiedzialna",
    "Decyzyjność / Czego wymaga by iść dalej",
    "Szacowany czas koordynacji (h)",
    "Uwagi"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Szerokości kolumn
col_widths = [5, 22, 35, 18, 16, 20, 14, 22, 45, 18, 35]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ============================================================
# DANE - Wszystkie działania z statusów
# ============================================================
data = [
    # === URODZINY LU ===
    ["URODZINY LU (3. URODZINY)", None, None, None, None, None, None, None, None, None, None],
    
    [1, "Urodziny LU", "Prezentacja konceptu urodzinowego", "Zrealizowano", "", "Do 23.06 (next status)", "", "Agata / Zespół", "Akceptacja konceptu przez Klienta, ustalenie elementów głównych (drony, lasery, fundusz)", 4, "Koncept omawiany na statusach 12.06 i 16.06"],
    
    [2, "Urodziny LU", "Wycena wydarzenia (drony, lasery, fundusz, gadżety)", "W trakcie", "", "Koniec czerwca 2026", "", "Agata", "Zebranie ofert od podwykonawców, wycena gadżetów (kubki ~150szt, katalogi), wycena pokazu dronów. Wycena musi zawierać koszty + ekwiwalent reklamowy/mediowy/PR", 8, "Zapytania gadżetowe wysłane, odpowiedzi z końcem kolejnego tygodnia"],
    
    [3, "Urodziny LU", "Weryfikacja terminu 7 września", "Do realizacji", "", "Do 23.06", "", "Agata / Zespół", "Sprawdzenie czy 7 września jest realny, czy trzeba przesunąć", 2, ""],
    
    [4, "Urodziny LU", "Zaproszenie / Save the date", "Do realizacji", "", "Wysyłka ~26 czerwca", "", "Zespół (Agata/Gosia)", "Gotowy koncept graficzny (wynikowa konceptu), akceptacja treści. Wysyłka osobiście (personalizowana)", 6, "Pierwsze wrazenie - nie spieszyć się. Poczekaj do końca miesiąca by zawrzeć więcej info"],
    
    [5, "Urodziny LU", "Komunikacja eventu (maile)", "Zaplanowano", "", "czerwiec-lipiec-sierpień 2026", "", "Zespół", "Gotowa lista odbiorców, treść mailingów", 6, "20 czerwek: 1st mailing save the date. Lipiec: przypominający. Sierpień: przypominający"],
    
    [6, "Urodziny LU", "Menu na wydarzenie", "Do realizacji", "", "Do 23.06 (omówienie)", "", "Marcin W (chef)", "Propozycja menu przez Marcina W, akceptacja Klienta", 4, "Marcin W ma przygotować propozycję do wtorkowego statusu"],
    
    [7, "Urodziny LU", "Scenariusz / Prowadzący", "Do realizacji", "", "Do końca czerwca", "", "Zespół", "Decyzja kto prowadzi (propozycja: Robert), ustalenie roli prowadzącego, scenariusz przebiegu", 8, "Rozważana rola Roberta - zna ludzi, może być naturalne. Paweł przemówi"],
    
    [8, "Urodziny LU", "Animacje na ekranach (diody, totemy, ekrany)", "Do realizacji", "", "Lipiec-sierpień", "", "Marcin C / Zespół", "Scenariusz animacji - historia LU + historia grupy + zakomunikowanie nowego miejsca", 10, "Wykorzystamy diode, totem i ekrany. Filmy i animacje opowiadające historię"],
    
    [9, "Urodziny LU", "Welcome / Good Night Pack", "Do realizacji", "", "Lipiec-sierpień", "", "Agata / Zespół", "Projekt gadżetu użytkowego z logo grupy (kubki termiczne?), decyzja o formie", 6, "Coś użytecznego z logo grupy, co będzie towarzyszyć gościom w pracy codziennie"],
    
    [10, "Urodziny LU", "Rozpoznanie partnerstwa (współpraca z właścicielem budynku)", "W trakcie", "", "Do końca czerwca", "", "Paweł", "Rozmowa z właścicielem budynku o partycypacji w kosztach (np. 100k przy pokazie dronów)", 4, "Paweł ma rozeznać temat na swoim poziomie kontaktów"],
    
    [11, "Urodziny LU", "Partnerstwo / Sponsorzy", "Do przemyślenia", "", "Lipiec", "", "Zespół", "Znalezienie partnera/media, który mógłby współfinansować (20k?). Decyzja o partnerstwie", 6, "Rozważacie jakieś partnerstwo? Klient pasujący, niekonkurujący"],
    
    [12, "Urodziny LU", "Budżet wydarzenia", "Do realizacji", "", "Do końca czerwca", "", "Zespół (Paweł)", "Policzenie kosztów, ustalenie możliwego budżetu", 4, "Potrzebny do przygotowania scenariusza, doboru prowadzącego, atrakcji"],

    # === MATERIAŁY REKLAMOWE LU ===
    ["MATERIAŁY REKLAMOWE LU", None, None, None, None, None, None, None, None, None, None],
    
    [13, "Materiały reklamowe LU", "Filmy reklamowe z sesji (Bartek, Gosia, Agata, Łukasz)", "Zrealizowano (akcept)", "", "Maj-czerwiec 2026", "", "Zespół kreatywny", "Akceptacja filmów przez Klienta (wycięcie 41 piętra, dokolorowanie)", 8, "Zaakceptowane. Wycinamy 41 piętro, dokolorowujemy elementy"],
    
    [14, "Materiały reklamowe LU", "Montaż materiałów (podwykonawca)", "Zrealizowano", "", "", "", "Podwykonawca (Agata koordynuje)", "Materiał źródłowy od Klienta", 2, "Agata montuje z podwykonawcą"],
    
    [15, "Materiały reklamowe LU", "Przygotowanie nowych formatów do kampanii Meta", "W trakcie", "", "Od kolejnego tygodnia po 12.06", "", "Zespół / Dom mediowy", "Gotowe filmy, akceptacja przed podmianką w kreacjach", 8, "Formaty reklamowe przygotowane pod Cele kampanii"],
    
    [16, "Materiały reklamowe LU", "Dźwięk do filmu z Gosią", "Do realizacji", "", "", "", "Zespół", "Dogranie dźwięku", 2, ""],

    # === NOWA FORMUŁA VIDEO / POSTY ===
    ["NOWA FORMUŁA VIDEO I POSTY", None, None, None, None, None, None, None, None, None, None],
    
    [17, "Nowa formuła video", "Zmiana formuły materiałów video do social media", "W trakcie", "", "Czerwiec 2026", "", "Michał / Marcin C / Marcin W", "Omówienie nowej formuły, przejrzenie materiału z sesji rok temu dla VF", 6, "Jutro (17.06) spotkanie Michała z Marcinem - przegląd materiału z sesji VF"],
    
    [18, "Nowa formuła video", "Scenariusze postów na konta", "W trakcie", "", "Czerwiec 2026", "", "Michał / Marcin C / Agata", "Treści postów do akceptacji, omówienie zmian w komunikacji", 6, "Zanim zmiany się zadzieją - posty idą na bieżąco. Omówić z Agatą"],
    
    [19, "Nowa formuła video", "Posty z sesji zeszłorocznej w willi", "W trakcie", "", "Czerwiec 2026", "", "Zespół", "Akceptacja treści i grafik", 4, "Materiał 'roboczy', zabawny, budujący atmosferę. Do akceptacji przed publikacją"],

    # === SESJA W VF ===
    ["SESJA W VILLA FOKSAL", None, None, None, None, None, None, None, None, None, None],
    
    [20, "Sesja VF", "Sesja zdjęciowa - ogród i kuchnia (24 czerwca)", "Zaplanowano", "", "24 czerwca 2026", "", "Marcin C / Marcin W", "Scenariusz ujęć od Marcina C dla Marcina W, potwierdzenie terminu", 6, "Marcin C przygotowuje szczegółowy scenariusz ujęć na cały dzień"],
    
    [21, "Sesja VF", "Materiał tour LU wewnętrzny", "Zrealizowano", "", "", "", "Zespół", "Oddany do dyspozycji Klienta", 2, "Dla rozmów z agencjami"],

    # === KAMPANIE REKLAMOWE ===
    ["KAMPANIE REKLAMOWE I MEDIA", None, None, None, None, None, None, None, None, None, None],
    
    [22, "Kampanie reklamowe", "Media plan 2026 - rewizja i korekty", "W trakcie", "", "Czerwiec 2026", "", "Zespół / Dom mediowy", "Przedstawienie proponowanych zmian/poszerzenia TG (wtorek po 16.06)", 8, "100k zł netto z obsługą mediową. Korekty w ramach budżetu rocznego"],
    
    [23, "Kampanie reklamowe", "Kampania Catering ( cały rok)", "W trakcie", "", "Styczeń-Grudzień 2026", "", "Dom mediowy", "Miesięczna optymalizacja", 2, "Ciągła kampania miesiąc w miesiąc"],
    
    [24, "Kampanie reklamowe", "Kampania Wesela (cały rok)", "W trakcie", "", "Styczeń-Grudzień 2026", "", "Dom mediowy", "Miesięczna optymalizacja", 2, "Ciągła kampania miesiąc w miesiąc"],
    
    [25, "Kampanie reklamowe", "Konferencje w LU (maj-czerwiec)", "W trakcie", "", "Maj-Czerwiec 2026", "", "Dom mediowy", "Włączone z końcem kwietnia", 2, "Zastąpiły 'małe konferencje'"],
    
    [26, "Kampanie reklamowe", "Garden Party VF (maj-czerwiec)", "W trakcie", "", "Maj-Czerwiec 2026", "", "Dom mediowy", "Przedłużone o dwa miesiące", 2, ""],
    
    [27, "Kampanie reklamowe", "Weryfikacja działań / zmiana podejścia / poszerzenie TG", "W trakcie", "", "Czerwiec 2026", "", "Dom mediowy / Zespół", "Przedstawienie propozycji zmian, oczekiwanych efektów", 4, "Spotkanie po statusie 16.06"],
    
    [28, "Kampanie reklamowe", "Reklamy z Robertem, Bartkiem, Łukaszem, Gosią, Agatą", "W trakcie", "", "", "", "Podwykonawca / Agata", "Gdzie są materiały? Kiedy będziemy mieć materiał?", 2, "Materiały montowane przez podwykonawcę"],

    # === NAGRYWKI / EVENTY ===
    ["NAGRYWKI I EVENTY", None, None, None, None, None, None, None, None, None, None],
    
    [29, "Nagrywki", "Wina gruzińskie (11 czerwca, 12-15:00)", "Zrealizowano", "11.06.2026", "11.06.2026", "1 dzień", "Zespół", "Materiał do montażu po ustaleniach Michała z Marcinem", 2, "Obszerna relacja. Ogród i oranżeria. Wszyscy managerowie"],
    
    [30, "Nagrywki", "Konkurs gotowania dziennikarzy (15 czerwca, 17:00)", "Zrealizowano", "15.06.2026", "15.06.2026", "1 dzień", "Robert / Zespół", "Materiał do social media i PR", 4, "Start nagrywek 18:00. Flowersy. Dziennikarze gotują 300 osób. Ela i Mateusz. Materiał poszedł w Pytanie na Śniadanie"],
    
    [31, "Nagrywki", "Koncert Monika Borzym jazz (19 czerwca, 18:00)", "Zaplanowano", "", "19 czerwca 2026", "1 dzień", "Zespół", "Nagranie materiału, serwowane menu", 4, ""],
    
    [32, "Nagrywki", "Gala kosmetyków (25 czerwca, 18:30/19:00)", "Zaplanowano", "", "25 czerwca 2026", "1 dzień", "Zespół", "Nagranie materiału", 3, "Wjazd scena o 18, start 19:00"],
    
    [33, "Nagrywki", "Catering wyjazdowy LU - Hebe + kosmetyki koreańskie (23 czerwca)", "Do potwierdzenia", "", "23 czerwca 2026", "1 dzień", "Zespół", "Potwierdzenie zmian", 2, "Proszę dać znać jak coś się zmieni"],
    
    [34, "Nagrywki", "Bionic premiera (11 maja)", "Do potwierdzenia", "11.05.2026", "11.05.2026", "1 dzień", "Agata", "Potwierdzenie kręcenia", 2, "Wstępnie 11 maja"],
    
    [35, "Nagrywki", "Wesele w LU (30 maja)", "Zaplanowano", "", "30.05.2026", "1 dzień", "Zespół", "Nagranie materiału", 3, "Wstępnie"],
    
    [36, "Nagrywki", "Manager - logistyka, wyporność, elastyczność (28 maja)", "Zrealizowano", "28.05.2026", "28.05.2026", "1 dzień", "Manager", "Nagranie wypowiedzi", 2, "Dwa cateringi + impreza VF oraz jeden w LU"],

    # === MATERIAŁ OPROWADZANIA PO LU ===
    ["MATERIAŁ OPROWADZANIA PO LU", None, None, None, None, None, None, None, None, None, None],
    
    [37, "Materiał oprowadzania", "Film oprowadzania po LU w trakcie prac", "W trakcie", "", "", "", "Zespół", "Dopracowanie materiału, udostępnienie na linku", 4, "Szybko okazał się wewnętrznie potrzebny. Klient z Krakowa obejrzał od razu. Gotowy do wysyłania Klientom/Agencjom"],

    # === CRM ===
    ["CRM - WDROŻENIE", None, None, None, None, None, None, None, None, None, None],
    
    [38, "CRM", "Warsztaty wdrożeniowe CRM (6 lipca + 13 lipca)", "Zaplanowano", "", "6 i 13 lipca 2026", "2 spotkania", "Paweł / Marcin C / Zespół", "Specyfikacja wymagań, wybór rozwiązania (HubSpot?), integracja z food costem, format ofert", 16, "Terminy w kalendarzu. Jedno spotkanie może być za mało"],
    
    [39, "CRM", "Definicja wymagań i specyfikacja", "Do realizacji", "", "Przed 6 lipca", "", "Zespół", "Ustalenie: głębokość wpięcia w CRM, integracja z food costem, format ofert, przepływ zleceń na kuchnię", 8, "Stworzyć dokument specyfikacji przed warsztatami"],

    # === STRONA WWW GRUPY ===
    ["STRONA WWW GRUPY", None, None, None, None, None, None, None, None, None, None],
    
    [40, "Strona www", "Strona parasolowa VillaFoksalGroup.pl", "W trakcie", "", "Czerwiec-Lipiec 2026", "", "Marcin C", "Projekt statyczny strony głównej + podstrony, struktura, copy, akceptacja Klienta", 12, "Marcin C przygotowuje projekt na wtorkowy status (17.06). Landing page: o nas, zakres, firmy w grupie"],
    
    [41, "Strona www", "Statyczne projekty stron LU i VF", "W trakcie", "", "Czerwiec 2026", "", "Marcin C", "Projekt graficzny, założenia struktury, copy", 8, "Do zderzenia z wiedzą i doświadczeniem zespołu"],
    
    [42, "Strona www", "Strona główna LU (projekt)", "W trakcie", "", "Czerwiec 2026", "", "Marcin C", "Akceptacja projektu", 6, ""],

    # === LOGO GRUPY ===
    ["LOGO GRUPY", None, None, None, None, None, None, None, None, None, None],
    
    [43, "Logo grupy", "Logo grupy - projekt finalny", "W trakcie", "", "Do 17.06 (wtorek)", "", "Zespół kreatywny", "Akceptacja Klienta, przełożenie na formaty", 6, "Wybraliśmy kierunek, we wtorek finalne + formaty"],
    
    [44, "Logo grupy", "Logo na materiałach i www", "Zaplanowano", "", "Po akceptacji logo", "", "Marcin C", "Gotowe logo w formatach wektorowych", 4, "Pokazanie jak logo pracuje na materiałach, www"],

    # === WIZYTÓWKI ===
    ["WIZYTÓWKI", None, None, None, None, None, None, None, None, None, None],
    
    [45, "Wizytówki", "Wizytówki - wycena i druk", "W trakcie", "", "Do końca czerwca (realizacja: miesiąc)", "", "Agata / Zespół", "Decyzja: stanowiska na wizytówkach (Agata vs Robert/Gosia - gradacja). Pliki do drukarni", 6, "Wybrane kartony - dłuższy czas oczekiwania. Alternatywa: jednokolorowy karton (tańszy)"],
    
    [46, "Wizytówki", "Wizytówki - alternatywna wersja (jednokolorowy karton)", "Do realizacji", "", "", "", "Agata", "Wycena obu wersji, decyzja Klienta", 2, "Koszt jednostkowy niższy, czas oczekiwania na karton podobny"],

    # === DOMENY ===
    ["DOMENY", None, None, None, None, None, None, None, None, None, None],
    
    [47, "Domeny", "Zakup domen dodatkowych", "W trakcie", "", "", "", "Marcin C", "Decyzja Klienta jakie domeny dodatkowe (poza głównymi)", 2, "VillaFoksalGroup.pl, GrupaVillaFoksal.pl, VillaFoksalgrupa.pl, GrupaFoksal.pl"],
    
    [48, "Domeny", "Domena VFG.pl (premium)", "Do decyzji", "", "", "", "Klient", "Czy kupować za premium?", 1, "Premium.pl - do kupienia od kogoś. Do decyzji"],

    # === PRACODAWCY RP ===
    ["PRACODAWCY RP", None, None, None, None, None, None, None, None, None, None],
    
    [49, "Pracodawcy RP", "Śniadanie u Pracodawcy RP", "Zrealizowano", "", "18-19.05.2026", "1 dzień", "Paweł / Zespół", "Prezentacja grupy (LU, VF, CVF, grupa)", 2, "Paweł mówi w kilku słowach o poszczególnych podmiotach"],
    
    [50, "Pracodawcy RP", "Mailing w Pracodawcy RP", "Zrealizowano", "", "", "", "Zespół", "Przygotowanie treści do akceptu", 2, "Będziemy w mailingu"],
    
    [51, "Pracodawcy RP", "Materiał w miesięczniku Pracodawcy RP", "W trakcie", "", "", "", "Zespół", "Przygotowanie treści", 2, "Będziemy mieć materiał w miesięczniku"],

    # === FORBES ===
    ["FORBES", None, None, None, None, None, None, None, None, None, None],
    
    [52, "Forbes", "Współpraca z Forbes", "Oczekiwanie", "", "", "", "Zespół", "Oczekiwanie na ofertę współpracy", 1, "Jeszcze nie dał oferty"],

    # === STRATEGIA GRUPY ===
    ["STRATEGIA GRUPY - WARSZTATY", None, None, None, None, None, None, None, None, None, None],
    
    [53, "Warsztaty strategiczne", "Warsztat 1 - 29 czerwca (poniedziałek)", "Zaplanowano", "", "29 czerwca 2026", "1 dzień", "Paweł / Marcin C / Zespół", "Sprzedaż, managerowie po jednym z lokalu, Marcin C, bez kuchni", 8, "Strategia komunikacji, oferta, format, grupa docelowa, pozycjonowanie LU i VF, konkurencja"],
    
    [54, "Warsztaty strategiczne", "Warsztat 2 - 6 lipca (poniedziałek)", "Zaplanowano", "", "6 lipca 2026", "1 dzień", "Paweł / Marcin C / Zespół", "Kontynuacja prac", 8, "Cel: wypracowanie strategii podwaleiny do stron, SEO, komunikacji, spotkania z zespołem"],
    
    [55, "Warsztaty strategiczne", "Spotkanie 'Kim jesteśmy na rynku' (dla zespołu)", "Zaplanowano", "", "Lipiec-Sierpień 2026", "", "Paweł / Zespół", "Gotowa strategia z warsztatów", 4, "Komunikacja wewnętrzna - całej strukturze pracowników"],

    # === IDENTYFIKACJA GRUPY ===
    ["IDENTYFIKACJA GRUPY", None, None, None, None, None, None, None, None, None, None],
    
    [56, "Identyfikacja grupy", "Zmiana firmy na Villa Foksal Group", "W trakcie", "", "", "", "Zespół / Marcin C", "Wdrożenie zmian w materiałach, komunikacji", 8, "Omawiane na spotkaniu 11.05"],
    
    [57, "Identyfikacja grupy", "Narracja grupy - 'Gościnność zapisana w historii miejsca'", "W trakcie", "", "", "", "Zespół", "Dopracowanie treści, wdrożenie na strony, mailingi, popupy, komunikacja zewnętrzna/wewnętrzna", 10, "Bazowa treść gotowa. Historia od XVIII w. przez Vauxhall, Zamoyskich, Café Bodo po dziś"],
    
    [58, "Identyfikacja grupy", "Sąsiad - Pałac na Foksal", "Zamknięto", "", "", "", "Zespół", "Brak zagrożenia - inne miejsce, inny adres", 0, "Sprawdzone - kompletnie inne miejsce. Brak działań"],

    # === INNE ===
    ["INNE", None, None, None, None, None, None, None, None, None, None],
    
    [59, "Inne", "Targi ślubne wrzesień - after party w LU", "Do sprawdzenia", "", "", "", "Zespół", "Sprawdzenie czy możliwe i na jakich warunkach", 2, "Pomysł - czy after party targów ślubnych w LU"],
    
    [60, "Inne", "Dzień otwarty: partnerzy i agencje (luty 2027)", "Do przemyślenia", "", "Luty 2027", "", "Zespół", "Decyzja czy robimy, koncepcja", 2, "Jak w 2025"],
    
    [61, "Inne", "Event w ogrodzie VF (marzec/kwiecień 2027)", "Do przemyślenia", "", "Marzec-Kwiecień 2027", "", "Zespół", "Decyzja", 1, ""],
    
    [62, "Inne", "SEO i strony www", "W trakcie", "", "Lipiec-Sierpień 2026", "", "Marcin C / Zespół", "Budowa SEO oparta o strategię z warsztatów", 10, "Wdrażanie po wypracowaniu strategii w czerwcu"],
    
    [63, "Inne", "Montaż Jägermeister - skrócenie do 45s", "Do realizacji", "", "", "", "Zespół", "Skrócenie + zmniejszenie nasycenia kolorów", 1, ""],
]

# Wypełnianie danych
row_num = 2
for item in data:
    # Sprawdź czy to nagłówek kategorii
    if item[1] is None:
        # Nagłówek kategorii
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=11)
        cell = ws.cell(row=row_num, column=1, value=item[0])
        cell.font = category_font
        cell.fill = category_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
        for col in range(2, 12):
            ws.cell(row=row_num, column=col).fill = category_fill
            ws.cell(row=row_num, column=col).border = thin_border
    else:
        # Normalny wiersz danych
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Kolorowanie statusów
            if col_idx == 4:  # Kolumna statusu
                if value == "Zrealizowano":
                    cell.fill = status_realized
                elif value in ["W trakcie", "Do potwierdzenia"]:
                    cell.fill = status_progress
                elif value in ["Zaplanowano", "Do realizacji", "Do przemyślenia", "Do sprawdzenia", "Do decyzji"]:
                    cell.fill = status_planned
                elif value == "Zablokowano":
                    cell.fill = status_blocked
            
            # Formatowanie liczb w kolumnie godzin
            if col_idx == 10 and isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_num += 1

# Zamrożenie pierwszego wiersza
ws.freeze_panes = 'A2'

# ============================================================
# SHEET 2: PODSUMOWANIE WEDŁUG STATUSÓW
# ============================================================
ws2 = wb.create_sheet("Podsumowanie wg Statusów")

# Nagłówki
podsumowanie_headers = ["Status", "Liczba działań", "Szacowany łączny czas koordynacji (h)"]
for col, header in enumerate(podsumowanie_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 35

# Zliczanie według statusów
status_count = {}
status_hours = {}
for item in data:
    if item[1] is not None:  # Nie nagłówek kategorii
        status = item[3]
        hours = item[9] if isinstance(item[9], (int, float)) else 0
        if status not in status_count:
            status_count[status] = 0
            status_hours[status] = 0
        status_count[status] += 1
        status_hours[status] += hours

status_order = ["Zrealizowano", "W trakcie", "Zaplanowano", "Do realizacji", "Do potwierdzenia", 
                "Do przemyślenia", "Do sprawdzenia", "Do decyzji", "Oczekiwanie", "Zablokowano"]

row_num = 2
for status in status_order:
    if status in status_count:
        ws2.cell(row=row_num, column=1, value=status).border = thin_border
        ws2.cell(row=row_num, column=2, value=status_count[status]).border = thin_border
        ws2.cell(row=row_num, column=3, value=status_hours[status]).border = thin_border
        
        # Kolorowanie
        cell_status = ws2.cell(row=row_num, column=1)
        if status == "Zrealizowano":
            cell_status.fill = status_realized
        elif status in ["W trakcie", "Do potwierdzenia"]:
            cell_status.fill = status_progress
        elif status in ["Zaplanowano", "Do realizacji", "Do przemyślenia", "Do sprawdzenia", "Do decyzji"]:
            cell_status.fill = status_planned
        elif status == "Zablokowano":
            cell_status.fill = status_blocked
        
        row_num += 1

# Wiersz podsumowania
ws2.cell(row=row_num + 1, column=1, value="RAZEM").font = bold_font
ws2.cell(row=row_num + 1, column=2, value=sum(status_count.values())).font = bold_font
ws2.cell(row=row_num + 1, column=3, value=sum(status_hours.values())).font = bold_font

# ============================================================
# SHEET 3: PODSUMOWANIE WEDŁUG KATEGORII
# ============================================================
ws3 = wb.create_sheet("Podsumowanie wg Kategorii")

kategoria_headers = ["Kategoria", "Liczba działań", "Szacowany czas koordynacji (h)"]
for col, header in enumerate(kategoria_headers, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 35

# Zliczanie według kategorii
kategoria_count = {}
kategoria_hours = {}
for item in data:
    if item[1] is not None:  # Nie nagłówek kategorii
        kategoria = item[1]
        hours = item[9] if isinstance(item[9], (int, float)) else 0
        if kategoria not in kategoria_count:
            kategoria_count[kategoria] = 0
            kategoria_hours[kategoria] = 0
        kategoria_count[kategoria] += 1
        kategoria_hours[kategoria] += hours

row_num = 2
for kategoria in sorted(kategoria_count.keys()):
    ws3.cell(row=row_num, column=1, value=kategoria).border = thin_border
    ws3.cell(row=row_num, column=2, value=kategoria_count[kategoria]).border = thin_border
    ws3.cell(row=row_num, column=3, value=kategoria_hours[kategoria]).border = thin_border
    row_num += 1

ws3.cell(row=row_num + 1, column=1, value="RAZEM").font = bold_font
ws3.cell(row=row_num + 1, column=2, value=sum(kategoria_count.values())).font = bold_font
ws3.cell(row=row_num + 1, column=3, value=sum(kategoria_hours.values())).font = bold_font

# ============================================================
# SHEET 4: HARMONOGRAM CZASOWY
# ============================================================
ws4 = wb.create_sheet("Harmonogram Czasowy")

harmonogram_headers = ["L.p.", "Temat / Działanie", "Kategoria", "Status", "Termin", "Priorytet czasowy"]
for col, header in enumerate(harmonogram_headers, 1):
    cell = ws4.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 40
ws4.column_dimensions['C'].width = 22
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 25
ws4.column_dimensions['F'].width = 20

# Działania posegregowane chronologicznie
harmonogram_data = [
    # Natychmiastowe / czerwiec 2026
    [1, "Wycena wydarzenia urodzinowego (drony, lasery, gadżety)", "Urodziny LU", "W trakcie", "Koniec czerwca 2026", "WYSOKI"],
    [2, "Zaproszenie / Save the date (wysyłka 26.06)", "Urodziny LU", "Do realizacji", "26 czerwca 2026", "WYSOKI"],
    [3, "Menu na urodziny (propozycja Marcina W)", "Urodziny LU", "Do realizacji", "23 czerwca 2026", "WYSOKI"],
    [4, "Sesja zdjęciowa VF - ogród i kuchnia", "Sesja VF", "Zaplanowano", "24 czerwca 2026", "WYSOKI"],
    [5, "Scenariusz ujęć do sesji VF (Marcin C)", "Sesja VF", "W trakcie", "Do 23 czerwca 2026", "WYSOKI"],
    [6, "Przedstawienie zmian TG (dom mediowy)", "Kampanie reklamowe", "W trakcie", "17-18 czerwca 2026", "WYSOKI"],
    [7, "Nowa formuła video - spotkanie Michał/Marcin", "Nowa formuła video", "W trakcie", "17 czerwca 2026", "WYSOKI"],
    [8, "Projekt statyczny strony www (Marcin C)", "Strona www", "W trakcie", "17 czerwca 2026", "WYSOKI"],
    [9, "Logo grupy - pokazanie finalne", "Logo grupy", "W trakcie", "17 czerwca 2026", "WYSOKI"],
    [10, "Koncert Monika Borzym", "Nagrywki", "Zaplanowano", "19 czerwca 2026", "ŚREDNI"],
    [11, "Catering Hebe + kosmetyki koreańskie", "Nagrywki", "Do potwierdzenia", "23 czerwca 2026", "ŚREDNI"],
    [12, "Gala kosmetyków", "Nagrywki", "Zaplanowano", "25 czerwca 2026", "ŚREDNI"],
    [13, "Warsztat strategiczny 1", "Warsztaty strategiczne", "Zaplanowano", "29 czerwca 2026", "WYSOKI"],
    [14, "Wycena gadżetów - odpowiedzi", "Urodziny LU", "W trakcie", "Koniec czerwca 2026", "ŚREDNI"],
    [15, "Scenariusz prowadzącego urodziny", "Urodziny LU", "Do realizacji", "Koniec czerwca 2026", "ŚREDNI"],
    
    # Lipiec 2026
    [16, "Warsztat strategiczny 2", "Warsztaty strategiczne", "Zaplanowano", "6 lipca 2026", "WYSOKI"],
    [17, "Warsztaty CRM (6 lipca)", "CRM", "Zaplanowano", "6 lipca 2026", "WYSOKI"],
    [18, "Warsztaty CRM (13 lipca)", "CRM", "Zaplanowano", "13 lipca 2026", "WYSOKI"],
    [19, "Mailing przypominający urodziny", "Urodziny LU", "Zaplanowano", "Lipiec 2026", "ŚREDNI"],
    [20, "Animacje na ekranach (historia LU + grupa)", "Urodziny LU", "Do realizacji", "Lipiec 2026", "ŚREDNI"],
    [21, "Welcome/Good Night Pack", "Urodziny LU", "Do realizacji", "Lipiec 2026", "ŚREDNI"],
    [22, "Budowa stron www (po strategii)", "Strona www", "Zaplanowano", "Lipiec-Sierpień 2026", "WYSOKI"],
    [23, "SEO (po strategii)", "Inne", "W trakcie", "Lipiec-Sierpień 2026", "WYSOKI"],
    
    # Sierpień 2026
    [24, "Mailing przypominający urodziny", "Urodziny LU", "Zaplanowano", "Sierpień 2026", "ŚREDNI"],
    [25, "Spotkanie 'Kim jesteśmy na rynku' (zespoł)", "Warsztaty strategiczne", "Zaplanowano", "Sierpień 2026", "ŚREDNI"],
    
    # Wrzesień 2026
    [26, "3 Urodziny LU - wydarzenie!", "Urodziny LU", "Zaplanowano", "7 września 2026", "WYSOKI"],
    [27, "Targi ślubne + after party w LU", "Inne", "Do sprawdzenia", "Wrzesień 2026", "NISKI"],
]

for row_idx, item in enumerate(harmonogram_data, 2):
    for col_idx, value in enumerate(item, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Kolorowanie priorytetu
        if col_idx == 6:
            if value == "WYSOKI":
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            elif value == "ŚREDNI":
                cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                cell.font = Font(name="Calibri", bold=True, size=10)
            elif value == "NISKI":
                cell.fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
                cell.font = Font(name="Calibri", bold=True, size=10)

ws4.freeze_panes = 'A2'

# ============================================================
# SHEET 5: POTRZEBNE DECYZJE
# ============================================================
ws5 = wb.create_sheet("Potrzebne Decyzje")

decyzje_headers = ["L.p.", "Temat", "Od kogo", "Termin", "Status", "Wpływ na dalsze działania"]
for col, header in enumerate(decyzje_headers, 1):
    cell = ws5.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 40
ws5.column_dimensions['C'].width = 20
ws5.column_dimensions['D'].width = 20
ws5.column_dimensions['E'].width = 15
ws5.column_dimensions['F'].width = 50

decyzje_data = [
    [1, "Czy 7 września jest realnym terminem na urodziny?", "Zespół", "Do 23 czerwca", "Oczekuje", "Wpływa na harmonogram komunikacji i przygotowań"],
    [2, "Kto prowadzi urodziny? (propozycja: Robert)", "Robert / Zespół", "Koniec czerwca", "Oczekuje", "Scenariusz wydarzenia, rola prowadzącego"],
    [3, "Czy właściciele budynku partycypują w kosztach (100k)?", "Paweł", "Do końca czerwca", "W trakcie", "Wpływa na budżet i zakres atrakcji (drony, lasery)"],
    [4, "Budżet na urodziny - ile możemy wydać?", "Paweł / Zespół", "Do końca czerwca", "Oczekuje", "Skala wydarzenia, atrakcje, gadżety"],
    [5, "Czy robimy partnerstwo/sponsoring z mediami?", "Zespół", "Lipiec", "Do przemyślenia", "Dodatkowe środki (20k?) lub promocja"],
    [6, "Jaka forma wizytówek - stanowiska na wizytówkach?", "Klient", "Do końca czerwca", "Oczekuje", "Projekt graficzny wizytówek"],
    [7, "Jakie domeny dodatkowe kupujemy?", "Klient", "Wkrótce", "W trakcie", "Bezpieczeństwo domen, przekierowania"],
    [8, "Czy kupujemy VFG.pl (premium)?", "Klient", "Wkrótce", "Do decyzji", "Alternatywna domena"],
    [9, "Jaki system CRM wybieramy?", "Zespół", "Przed 6 lipca", "Oczekuje", "Wdrożenie CRM, integracja z food costem"],
    [10, "Czy after party targów ślubnych w LU?", "Zespół", "Wkrótce", "Do sprawdzenia", "Dodatkowy event we wrześniu"],
    [11, "Akceptacja projektu statycznego strony www", "Klient", "Do końca czerwca", "Oczekuje", "Budowa stron, SEO, komunikacja"],
    [12, "Akceptacja logo grupy", "Klient", "17 czerwca", "Oczekuje", "Wszystkie materiały, strony, wizytówki"],
    [13, "Czy cateringt Hebe 23 czerwca się odbywa?", "Zespół", "Do 20 czerwca", "Oczekuje", "Nagranie materiału"],
    [14, "Scenariusz ujęć do sesji VF 24 czerwca", "Marcin C", "Do 23 czerwca", "W trakcie", "Realizacja sesji"],
    [15, "Treści postów - akceptacja przed publikacją", "Marcin C / Agata", "Na bieżąco", "W trakcie", "Komunikacja w social media"],
]

for row_idx, item in enumerate(decyzje_data, 2):
    for col_idx, value in enumerate(item, 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws5.freeze_panes = 'A2'

# Zapisanie pliku
output_path = r"C:\Users\Malgorzata\Desktop\GRUPA VILLA FOKSAL ANTYGRAVITY\GVF\GVF_Zestawienie_Dzialan.xlsx"
wb.save(output_path)
print(f"Plik zapisany: {output_path}")
print(f"Liczba działań: {sum(status_count.values())}")
print(f"Łączny szacowany czas koordynacji: {sum(status_hours.values())}h")
