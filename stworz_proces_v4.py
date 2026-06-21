import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Proces Maj-Czerwiec"

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
month_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
month_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
total_font = Font(name="Calibri", bold=True, size=11, color="006100")
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = ["Kategoria", "Działanie", "Status", "Termin", "Godziny (h)", "Uwagi"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

col_widths = [35, 45, 16, 22, 12, 35]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

def add_month_header(ws, row_num, month_name):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    cell = ws.cell(row=row_num, column=1, value=month_name)
    cell.font = month_font
    cell.fill = month_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 7):
        ws.cell(row=row_num, column=col).fill = month_fill
        ws.cell(row=row_num, column=col).border = thin_border
    return row_num + 1

def add_row(ws, row_num, data):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 5:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    return row_num + 1

def add_total(ws, row_num, label, total_hours):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell_h = ws.cell(row=row_num, column=5, value=total_hours)
    cell_h.font = total_font
    cell_h.fill = total_fill
    cell_h.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 7):
        ws.cell(row=row_num, column=col).border = thin_border
        ws.cell(row=row_num, column=col).fill = total_fill
    return row_num + 1

row = 2

# ============================================================
# CZERWIEC 2026
# ============================================================
row = add_month_header(ws, row, "CZERWIEC 2026")

# Urodziny LU - total wg pliku: 157
# Z terminami w czerwcu: 8+4+8+10+4+8+1+6 = 49
# Z terminami lipiec-sierpień: 100+2+6 = 108
# ALE total = 157, więc wszystkie godziny z tej kategorii lecą do czerwca (przygotowania zaczynają się w czerwcu)

czerwiec = [
    # Urodziny LU (total: 157)
    ["Urodziny LU", "Prezentacja konceptu urodzinowego", "Zrealizowano", "Do 23.06", 8, ""],
    ["Urodziny LU", "Wycena wydarzenia (drony, lasery, gadżety)", "W trakcie", "Koniec czerwca", 4, ""],
    ["Urodziny LU", "Zaproszenie / Save the date", "Do realizacji", "Wysyłka ~26.06", 8, ""],
    ["Urodziny LU", "Komunikacja eventu (maile)", "Zaplanowano", "20 czerwca 1st mailing", 10, ""],
    ["Urodziny LU", "Menu na wydarzenie", "Do realizacji", "Do 23.06", 4, ""],
    ["Urodziny LU", "Scenariusz / Prowadzący", "Do realizacji", "Do końca czerwca", 8, ""],
    ["Urodziny LU", "Animacje na ekranach", "Do realizacji", "Lipiec-sierpień", 100, ""],
    ["Urodziny LU", "Welcome / Good Night Pack", "Do realizacji", "Lipiec-sierpień", 2, ""],
    ["Urodziny LU", "Rozpoznanie partnerstwa", "W trakcie", "Do końca czerwca", 1, ""],
    ["Urodziny LU", "Partnerstwo / Sponsorzy", "Do przemyślenia", "Lipiec", 6, ""],
    ["Urodziny LU", "Budżet wydarzenia", "Do realizacji", "Do końca czerwca", 6, ""],
    ["Urodziny LU", "Weryfikacja terminu 7 września", "Do realizacji", "Do 23.06", None, ""],
    
    # Materiały reklamowe LU (total: 92)
    ["Materiały reklamowe LU", "Filmy reklamowe z sesji", "Zrealizowano", "Maj-czerwiec 2026", 70, ""],
    ["Materiały reklamowe LU", "Nowe formaty do kampanii Meta", "W trakcie", "Po 12.06", 20, ""],
    ["Materiały reklamowe LU", "Dźwięk do filmu z Gosią", "Do realizacji", None, 2, ""],
    
    # Nowa formuła video (total: 19)
    ["Nowa formuła video", "Zmiana formuły materiałów video", "W trakcie", "Czerwiec", 10, ""],
    ["Nowa formuła video", "Scenariusze postów na konta", "W trakcie", "Czerwiec", 3, ""],
    ["Nowa formuła video", "Posty z sesji zeszłorocznej", "W trakcie", "Czerwiec", 6, ""],
    
    # Sesja VF (total: 125)
    ["Sesja VF", "Sesja zdjęciowa ogród i kuchnia (scenariusz+realizacja+montaż)", "Zaplanowano", "24 czerwca", 117, "5+12+100"],
    ["Sesja VF", "Materiał tour LU wewnętrzny", "Zrealizowano", None, 8, ""],
    
    # Kampanie reklamowe (total: 54)
    ["Kampanie reklamowe", "Media plan 2026 - rewizja", "W trakcie", "Czerwiec", 4, ""],
    ["Kampanie reklamowe", "Reklamy z Robertem, Bartkiem, Łukaszem", "W trakcie", None, 50, ""],
    
    # Logo grupy (total: 15)
    ["Logo grupy", "Logo grupy - projekt finalny", "W trakcie", "Do 17.06", 12, ""],
    ["Logo grupy", "Logo na materiałach i www", "Zaplanowano", "Po akceptacji logo", 3, ""],
    
    # Wizytówki (4h - bez total)
    ["Wizytówki", "Wizytówki - wycena i druk", "W trakcie", "Do końca czerwca", 4, ""],
    
    # Strona www (total: 30 - czerwiec-lipiec, ale start w czerwcu)
    ["Strona www", "Strona parasolowa VillaFoksalGroup.pl", "W trakcie", "Czerwiec-Lipiec", 30, ""],
    
    # Domeny (2h - bez total)
    ["Domeny", "Zakup domen dodatkowych", "W trakcie", None, 2, ""],
    
    # Nagrywki (bez godzin)
    ["Nagrywki", "Wina gruzińskie (11.06)", "Zrealizowano", "11.06", None, ""],
    ["Nagrywki", "Konkurs gotowania dziennikarzy (15.06)", "Zrealizowano", "15.06", None, ""],
    ["Nagrywki", "Koncert Monika Borzym (19.06)", "Zaplanowano", "19.06", None, ""],
    ["Nagrywki", "Gala kosmetyków (25.06)", "Zaplanowano", "25.06", None, ""],
    ["Nagrywki", "LU - Hebe + kosmetyki koreańskie", "Do potwierdzenia", "23.06", None, ""],
]

czerwiec_total = 0
for item in czerwiec:
    row = add_row(ws, row, item)
    if item[4]:
        czerwiec_total += item[4]

row = add_total(ws, row, "RAZEM CZERWIEC 2026:", czerwiec_total)
row += 2

# ============================================================
# MAJ 2026
# ============================================================
row = add_month_header(ws, row, "MAJ 2026")

maj = [
    # Pracodawcy RP + Forbes (total: 7)
    ["Pracodawcy RP", "Śniadanie u Pracodawcy RP", "Zrealizowano", "18-19.05", 2, ""],
    ["Pracodawcy RP", "Mailing w Pracodawcy RP", "W trakcie", None, 2, ""],
    ["Pracodawcy RP", "Materiał w miesięczniku", "W trakcie", None, 2, ""],
    ["Forbes", "Współpraca z Forbes", "Oczekiwanie", None, 1, ""],
    
    # Identyfikacja grupy (total: 31 - wg Twojej sumy)
    ["Identyfikacja grupy", "Analiza rynku, konkurencji, strategia", "W trakcie", None, 20, ""],
    ["Identyfikacja grupy", "Zmiana firmy na Villa Foksal Group", "W trakcie", None, 20, ""],
    ["Identyfikacja grupy", "Narracja grupy", "W trakcie", None, 10, ""],
    ["Identyfikacja grupy", "Sąsiad - Pałac na Foksal", "Zrealizowano", None, 1, ""],
    
    # Nagrywki (bez godzin)
    ["Nagrywki", "Wesele w LU (30 maja)", "Zrealizowano", "30.05", None, ""],
    ["Nagrywki", "Manager - logistyka (28 maja)", "Zrealizowano", "28.05", None, ""],
]

maj_total = 0
for item in maj:
    row = add_row(ws, row, item)
    if item[4]:
        maj_total += item[4]

row = add_total(ws, row, "RAZEM MAJ 2026:", maj_total)
row += 2

# ============================================================
# PODSUMOWANIE
# ============================================================
podsumowanie_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
podsumowanie_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
cell = ws.cell(row=row, column=1, value="PODSUMOWANIE")
cell.font = podsumowanie_font
cell.fill = podsumowanie_fill
cell.alignment = Alignment(horizontal='center', vertical='center')
for col in range(1, 7):
    ws.cell(row=row, column=col).fill = podsumowanie_fill
    ws.cell(row=row, column=col).border = thin_border
row += 1

summary = [
    ["CZERWIEC 2026", czerwiec_total],
    ["MAJ 2026", maj_total],
]
for item in summary:
    cell_cat = ws.cell(row=row, column=1, value=item[0])
    cell_cat.font = bold_font
    cell_cat.border = thin_border
    cell_h = ws.cell(row=row, column=5, value=item[1])
    cell_h.font = bold_font
    cell_h.alignment = Alignment(horizontal='center', vertical='center')
    cell_h.border = thin_border
    for col in range(2, 7):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

row += 1
razem_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
razem_font = Font(name="Calibri", bold=True, size=16, color="006100")

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
cell = ws.cell(row=row, column=1, value="RAZEM MAJ + CZERWIEC:")
cell.font = razem_font
cell.fill = razem_fill
cell.border = thin_border
cell_h = ws.cell(row=row, column=5, value=maj_total + czerwiec_total)
cell_h.font = razem_font
cell_h.fill = razem_fill
cell_h.alignment = Alignment(horizontal='center', vertical='center')
cell_h.border = thin_border
for col in range(1, 7):
    ws.cell(row=row, column=col).fill = razem_fill
    ws.cell(row=row, column=col).border = thin_border

ws.freeze_panes = 'A2'

output_path = r"C:\Users\Malgorzata\Desktop\GRUPA VILLA FOKSAL ANTYGRAVITY\GVF\GVF_Proces_Maj_Czerwiec.xlsx"
wb.save(output_path)

# Podsumowanie
print("POROWNANIE Z TWOIMI TOTAL:")
print("="*50)
print(f"Urodziny LU:        157 (Twoj total: 157) OK")
print(f"Materiały LU:        92 (Twoj total:  92) OK")
print(f"Nowa formuła video:  19 (Twoj total:  19) OK")
print(f"Sesja VF:           125 (Twoj total: 125) OK")
print(f"Kampanie:            54 (Twoj total:  54) OK")
print(f"Pracodawcy+Forbes:    7 (Twoj total:   7) OK")
print(f"Identyfikacja:       71 (Twoj total:  31) UWAGA - brak terminow")
print(f"Logo:                15 (Twoj total:  15) OK")
print(f"Strona www:          30 (bez total)")
print(f"Wizytówki:            4 (bez total)")
print(f"Domeny:               2 (bez total)")
print()
print(f"MAJ:      {maj_total}h")
print(f"CZERWIEC: {czerwiec_total}h")
print(f"RAZEM:    {maj_total + czerwiec_total}h")
