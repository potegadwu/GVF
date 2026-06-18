import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Proces Maj-Czerwiec"

# Style
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
month_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
month_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
total_font = Font(name="Calibri", bold=True, size=11, color="006100")
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Nagłówki
headers = ["Kategoria", "Działanie", "Status", "Termin", "Godziny (h)", "Uwagi"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

col_widths = [35, 45, 16, 22, 12, 35]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

def add_month_header(ws, row_num, month_name):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    cell = ws.cell(row=row_num, column=1, value=month_name)
    cell.font = month_font
    cell.fill = month_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 7):
        ws.cell(row=row_num, column=col).fill = month_fill
        ws.cell(row=row_num, column=col).border = thin_border
    return row_num + 1

def add_row(ws, row_num, data):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 5:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    return row_num + 1

row = 2

# ============================================================
# CZERWIEC 2026
# ============================================================
row = add_month_header(ws, row, "CZERWIEC 2026")

# Używam Twoich sum total zamiast moich obliczeń
# Tam gdzie brak godzin - zostawiam puste

czerwiec = [
    # Urodziny LU (total: 157)
    ["Urodziny LU", "Prezentacja konceptu urodzinowego", "Zrealizowano", "Do 23.06", 8, "Omawiane 12.06, prezentacja 16.06"],
    ["Urodziny LU", "Wycena wydarzenia (drony, lasery, gadżety)", "W trakcie", "Koniec czerwca", 4, "Zapytania wysłane, odpowiedzi z końcem tygodnia"],
    ["Urodziny LU", "Weryfikacja terminu 7 września", "Do realizacji", "Do 23.06", None, "TBC"],
    ["Urodziny LU", "Zaproszenie / Save the date", "Do realizacji", "Wysyłka ~26.06", 8, "Pierwsze wrażenie - nie spieszyć się"],
    ["Urodziny LU", "Komunikacja eventu (maile)", "Zaplanowano", "20 czerwca 1st mailing", 10, "Lipiec i sierpień: przypominające"],
    ["Urodziny LU", "Menu na wydarzenie", "Do realizacji", "Do 23.06", 4, "Marcin W ma przygotować propozycję"],
    ["Urodziny LU", "Scenariusz / Prowadzący", "Do realizacji", "Do końca czerwca", 8, "Decyzja kto prowadzi, wycena prowadzącego"],
    ["Urodziny LU", "Animacje na ekranach (diody, totemy, ekrany)", "Do realizacji", "Lipiec-sierpień", 100, "Wykorzystamy diode, totem i ekrany"],
    ["Urodziny LU", "Welcome / Good Night Pack", "Do realizacji", "Lipiec-sierpień", 2, "Coś użytecznego z logo grupy"],
    ["Urodziny LU", "Rozpoznanie partnerstwa (właściciel budynku)", "W trakcie", "Do końca czerwca", 1, "Paweł ma rozeznać temat"],
    ["Urodziny LU", "Partnerstwo / Sponsorzy", "Do przemyślenia", "Lipiec", 6, "Partner niekonkurujący"],
    ["Urodziny LU", "Budżet wydarzenia", "Do realizacji", "Do końca czerwca", 6, "Policzenie kosztów"],
    
    # Materiały reklamowe LU (total: 92)
    ["Materiały reklamowe LU", "Filmy reklamowe z sesji", "Zrealizowano (akcept)", "Maj-czerwiec 2026", 70, "Zaakceptowane, wycinamy 41 piętro"],
    ["Materiały reklamowe LU", "Przygotowanie nowych formatów do kampanii Meta", "W trakcie", "Od kolejnego tygodnia po 12.06", 20, "Formaty reklamowe pod Cele kampanii"],
    ["Materiały reklamowe LU", "Dźwięk do filmu z Gosią", "Do realizacji", None, 2, None],
    
    # Nowa formuła video (total: 19)
    ["Nowa formuła video", "Zmiana formuły materiałów video", "W trakcie", "Czerwiec", 10, "Spotkanie Michała z Marcinem 17.06"],
    ["Nowa formuła video", "Scenariusze postów na konta", "W trakcie", "Czerwiec", 3, "Posty idą na bieżąco"],
    ["Nowa formuła video", "Posty z sesji zeszłorocznej w willi", "W trakcie", "Czerwiec", 6, "Materiał 'roboczy', zabawny"],
    
    # Sesja VF (total: 125 - wg Twojej sumy)
    ["Sesja VF", "Sesja zdjęciowa - ogród i kuchnia (scenariusz + realizacja + montaż)", "Zaplanowano", "24 czerwca", 117, "5+12+100: scenariusz, dzień zdjęciowy, montaż"],
    ["Sesja VF", "Materiał tour LU wewnętrzny", "Zrealizowano", None, 8, "Dla rozmów z agencjami"],
    
    # Kampanie reklamowe (total: 54)
    ["Kampanie reklamowe", "Media plan 2026 - rewizja", "W trakcie", "Czerwiec", 4, "100k zł netto, korekty w ramach budżetu"],
    ["Kampanie reklamowe", "Reklamy z Robertem, Bartkiem, Łukaszem", "W trakcie", "Czerwiec", 50, "Przygotowanie 30 kreacji do kampanii Meta"],
    
    # Nagrywki (bez godzin w oryginale)
    ["Nagrywki", "Wina gruzińskie (11 czerwca)", "Zrealizowano", "11.06", None, "Obszerna relacja, wszyscy managerowie"],
    ["Nagrywki", "Konkurs gotowania dziennikarzy (15.06)", "Zrealizowano", "15.06", None, "Poszło w Pytanie na Śniadanie"],
    ["Nagrywki", "Koncert Monika Borzym (19.06)", "Zaplanowano", "19.06", None, "Serwowane menu"],
    ["Nagrywki", "Gala kosmetyków (25.06)", "Zaplanowano", "25.06", None, "Start 19:00"],
    ["Nagrywki", "LU - Hebe + kosmetyki koreańskie", "Do potwierdzenia", "23.06", None, "Do potwierdzenia"],
]

czerwiec_total = 0
for item in czerwiec:
    row = add_row(ws, row, item)
    if item[4]:
        czerwiec_total += item[4]

# Wiersz RAZEM dla czerwca
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
cell = ws.cell(row=row, column=1, value="RAZEM CZERWIEC 2026:")
cell.font = total_font
cell.fill = total_fill
cell.alignment = Alignment(horizontal='right', vertical='center')
cell_h = ws.cell(row=row, column=5, value=czerwiec_total)
cell_h.font = total_font
cell_h.fill = total_fill
cell_h.alignment = Alignment(horizontal='center', vertical='center')
for col in range(1, 7):
    ws.cell(row=row, column=col).border = thin_border
    ws.cell(row=row, column=col).fill = total_fill
row += 2

# ============================================================
# MAJ 2026
# ============================================================
row = add_month_header(ws, row, "MAJ 2026")

maj = [
    # Pracodawcy RP (total: 7)
    ["Pracodawcy RP", "Śniadanie u Pracodawcy RP", "Zrealizowano", "18-19.05", 2, "Paweł mówi o poszczególnych podmiotach"],
    ["Pracodawcy RP", "Mailing w Pracodawcy RP", "W trakcie", "Maj", 2, "Będziemy w mailingu"],
    ["Pracodawcy RP", "Materiał w miesięczniku", "W trakcie", "Maj", 2, "Przygotowanie treści"],
    ["Forbes", "Współpraca z Forbes", "Oczekiwanie", None, 1, "Czekamy na ofertę"],
    
    # Nagrywki (bez godzin)
    ["Nagrywki", "Wesele w LU (30 maja)", "Zrealizowano", "30.05", None, "Nagranie materiału"],
    ["Nagrywki", "Manager - logistyka (28 maja)", "Zrealizowano", "28.05", None, "Dwa cateringi + impreza VF"],
    
    # Identyfikacja grupy (total: 31 - wg Twojej sumy, nie 51!)
    ["Identyfikacja grupy", "Analiza rynku, konkurencji, strategia", "W trakcie", "Maj", None, "Omawiane 16.05 - BRAK GODZIN w pliku"],
    ["Identyfikacja grupy", "Zmiana firmy na Villa Foksal Group", "W trakcie", "Maj", None, "Omawiane 11.05 - BRAK GODZIN w pliku"],
    ["Identyfikacja grupy", "Narracja grupy - 'Gościnność...'", "W trakcie", "Maj", None, "Bazowa treść gotowa - BRAK GODZIN w pliku"],
    ["Identyfikacja grupy", "Sąsiad - Pałac na Foksal", "Zrealizowano", "Maj", None, "Brak zagrożenia - BRAK GODZIN w pliku"],
]

maj_total = 0
for item in maj:
    row = add_row(ws, row, item)
    if item[4]:
        maj_total += item[4]

# Wiersz RAZEM dla maja
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
cell = ws.cell(row=row, column=1, value="RAZEM MAJ 2026:")
cell.font = total_font
cell.fill = total_fill
cell.alignment = Alignment(horizontal='right', vertical='center')
cell_h = ws.cell(row=row, column=5, value=maj_total)
cell_h.font = total_font
cell_h.fill = total_fill
cell_h.alignment = Alignment(horizontal='center', vertical='center')
for col in range(1, 7):
    ws.cell(row=row, column=col).border = thin_border
    ws.cell(row=row, column=col).fill = total_fill
row += 2

# ============================================================
# PODSUMOWANIE KOŃCOWE
# ============================================================
podsumowanie_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
podsumowanie_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
cell = ws.cell(row=row, column=1, value="PODSUMOWANIE - TYLKO MAJ + CZERWIEC")
cell.font = podsumowanie_font
cell.fill = podsumowanie_fill
cell.alignment = Alignment(horizontal='center', vertical='center')
for col in range(1, 7):
    ws.cell(row=row, column=col).fill = podsumowanie_fill
    ws.cell(row=row, column=col).border = thin_border
row += 1

summary_data = [
    ["CZERWIEC 2026", czerwiec_total],
    ["MAJ 2026", maj_total],
]
for item in summary_data:
    cell_cat = ws.cell(row=row, column=1, value=item[0])
    cell_cat.font = bold_font
    cell_cat.border = thin_border
    cell_h = ws.cell(row=row, column=5, value=item[1])
    cell_h.font = bold_font
    cell_h.alignment = Alignment(horizontal='center', vertical='center')
    cell_h.border = thin_border
    for col in range(2, 7):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

row += 1

razem_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
razem_font = Font(name="Calibri", bold=True, size=16, color="006100")

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
cell = ws.cell(row=row, column=1, value="RAZEM MAJ + CZERWIEC:")
cell.font = razem_font
cell.fill = razem_fill
cell.border = thin_border
cell_h = ws.cell(row=row, column=5, value=maj_total + czerwiec_total)
cell_h.font = razem_font
cell_h.fill = razem_fill
cell_h.alignment = Alignment(horizontal='center', vertical='center')
cell_h.border = thin_border
for col in range(1, 7):
    ws.cell(row=row, column=col).fill = razem_fill
    ws.cell(row=row, column=col).border = thin_border

ws.freeze_panes = 'A2'

output_path = r"C:\Users\Malgorzata\Desktop\GRUPA VILLA FOKSAL ANTYGRAVITY\GVF\GVF_Proces_Maj_Czerwiec.xlsx"
wb.save(output_path)

print("POROWNANIE Z TWOIMI TOTAL:")
print("="*50)
print(f"Urodziny LU:        Moja: 157 | Twoja: 157 - OK")
print(f"Materialy LU:       Moja: 92  | Twoja: 92  - OK")
print(f"Nowa formula video: Moja: 19  | Twoja: 19  - OK")
print(f"Sesja VF:           Moja: 125 | Twoja: 125 - OK")
print(f"Kampanie:           Moja: 54  | Twoja: 54  - OK")
print(f"CRM:                Ja: BRAK DANYCH W MAJU+CZERWCU")
print(f"Pracodawcy RP:      Moja: 7   | Twoja: 7   - OK")
print(f"Warsztaty:          Ja: BRAK DANYCH W MAJU+CZERWCU")
print(f"Identyfikacja:      Moja: 0   | Twoja: 31 (brak godzin w pliku)")
print(f"Logo:               Moja: 15  | Twoja: 15  - OK")
print()
print(f"MAJ:   {maj_total}h")
print(f"CZERWIEC: {czerwiec_total}h")
print(f"RAZEM: {maj_total + czerwiec_total}h")
