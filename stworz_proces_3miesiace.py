import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Proces Maj-Czerwiec-Lipiec"

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
month_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
month_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
total_font = Font(name="Calibri", bold=True, size=11, color="006100")
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = ["Kategoria", "Dzialanie", "Status", "Termin", "Godziny (h)", "Uwagi"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

col_widths = [35, 45, 16, 22, 12, 35]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

def add_month_header(ws, row_num, month_name):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    cell = ws.cell(row=row_num, column=1, value=month_name)
    cell.font = month_font
    cell.fill = month_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 7):
        ws.cell(row=row_num, column=col).fill = month_fill
        ws.cell(row=row_num, column=col).border = thin_border
    return row_num + 1

def add_row(ws, row_num, data):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 5:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    return row_num + 1

def add_total(ws, row_num, label, total_hours):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell_h = ws.cell(row=row_num, column=5, value=total_hours)
    cell_h.font = total_font
    cell_h.fill = total_fill
    cell_h.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 7):
        ws.cell(row=row_num, column=col).border = thin_border
        ws.cell(row=row_num, column=col).fill = total_fill
    return row_num + 1

row = 2

# ============================================================
# MAJ 2026
# ============================================================
row = add_month_header(ws, row, "MAJ 2026")

# Identyfikacja grupy - total 31, brak terminow - zakladam maj (omawiane 11.05 i 16.05)
# Materiały reklamowe - "maj/czerwiec" 70h -> 35h maj + 35h czerwiec
maj = [
    ["Pracodawcy RP", "Sniadanie u Pracodawcy RP", "Zrealizowano", "18-19.05", 2, ""],
    ["Pracodawcy RP", "Mailing w Pracodawcy RP", "W trakcie", None, 2, ""],
    ["Pracodawcy RP", "Material w miesieczniku", "W trakcie", None, 2, ""],
    ["Forbes", "Wspolpraca z Forbes", "Oczekiwanie", None, 1, ""],
    ["Nagrywki", "Wesele w LU (30 maja)", "Zrealizowano", "30.05", None, ""],
    ["Nagrywki", "Manager - logistyka (28 maja)", "Zrealizowano", "28.05", None, ""],
    ["Identyfikacja grupy", "Analiza rynku, konkurencji, strategia", "W trakcie", None, 20, ""],
    ["Identyfikacja grupy", "Zmiana firmy na Villa Foksal Group", "W trakcie", None, 20, ""],
    ["Identyfikacja grupy", "Narracja grupy", "W trakcie", None, 10, ""],
    ["Identyfikacja grupy", "Sasiad - Palac na Foksal", "Zrealizowano", None, 1, ""],
    # Materiały LU - połowa z "maj/czerwiec" 70h = 35h
    ["Materialy reklamowe LU", "Filmy reklamowe z sesji (polowa)", "Zrealizowano", "maj/czerwiec", 35, "Podzial: maj/czerwiec = 35+35"],
]

maj_total = 0
for item in maj:
    row = add_row(ws, row, item)
    if item[4]:
        maj_total += item[4]

row = add_total(ws, row, "RAZEM MAJ 2026:", maj_total)
row += 2

# ============================================================
# CZERWIEC 2026
# ============================================================
row = add_month_header(ws, row, "CZERWIEC 2026")

# Druga polowa materialow LU 70h = 35h
# Strona www "czerwiec-lipiec" 30h = 15+15
czerwiec = [
    # Urodziny LU - 157 total
    ["Urodziny LU", "Prezentacja konceptu", "Zrealizowano", "Do 23.06", 8, ""],
    ["Urodziny LU", "Wycena wydarzenia", "W trakcie", "Koniec czerwca", 4, ""],
    ["Urodziny LU", "Zaproszenie / Save the date", "Do realizacji", "Wysylka ~26.06", 8, ""],
    ["Urodziny LU", "Komunikacja eventu (maile)", "Zaplanowano", "20 czerwca 1st mailing", 10, ""],
    ["Urodziny LU", "Menu na wydarzenie", "Do realizacji", "Do 23.06", 4, ""],
    ["Urodziny LU", "Scenariusz / Prowadzacy", "Do realizacji", "Do konca czerwca", 8, ""],
    ["Urodziny LU", "Animacje na ekranach", "Do realizacji", "Lipiec-sierpien", 100, ""],
    ["Urodziny LU", "Welcome / Good Night Pack", "Do realizacji", "Lipiec-sierpien", 2, ""],
    ["Urodziny LU", "Rozpoznanie partnerstwa", "W trakcie", "Do konca czerwca", 1, ""],
    ["Urodziny LU", "Partnerstwo / Sponsorzy", "Do przemyslenia", "Lipiec", 6, ""],
    ["Urodziny LU", "Budzet wydarzenia", "Do realizacji", "Do konca czerwca", 6, ""],
    ["Urodziny LU", "Weryfikacja terminu 7 wrzesnia", "Do realizacji", "Do 23.06", None, ""],
    # Druga polowa materialow LU
    ["Materialy reklamowe LU", "Filmy reklamowe z sesji (polowa)", "Zrealizowano", "maj/czerwiec", 35, ""],
    ["Materialy reklamowe LU", "Nowe formaty do kampanii Meta", "W trakcie", "Po 12.06", 20, ""],
    ["Materialy reklamowe LU", "Dzwiek do filmu z Gosia", "Do realizacji", None, 2, ""],
    # Nowa formuza video - 19
    ["Nowa formula video", "Zmiana formuly materialow video", "W trakcie", "Czerwiec", 10, ""],
    ["Nowa formula video", "Scenariusze postow na konta", "W trakcie", "Czerwiec", 3, ""],
    ["Nowa formula video", "Posty z sesji zeszlorocznej", "W trakcie", "Czerwiec", 6, ""],
    # Sesja VF - 125
    ["Sesja VF", "Sesja zdjeciowa - ogród i kuchnia", "Zaplanowano", "24 czerwca", 117, "5+12+100"],
    ["Sesja VF", "Material tour LU wewnetrzny", "Zrealizowano", None, 8, ""],
    # Kampanie - 54
    ["Kampanie reklamowe", "Media plan 2026 - rewizja", "W trakcie", "Czerwiec", 4, ""],
    ["Kampanie reklamowe", "Reklamy z Robertem, Bartkiem, Lukaszem", "W trakcie", None, 50, ""],
    # Logo - 15
    ["Logo grupy", "Logo grupy - projekt finalny", "W trakcie", "Do 17.06", 12, ""],
    ["Logo grupy", "Logo na materialach i www", "Zaplanowano", "Po akceptacji logo", 3, ""],
    # Wizytowki - 4
    ["Wizytowki", "Wizytowki - wycena i druk", "W trakcie", "Do konca czerwca", 4, ""],
    # Strona www - polowa 30h = 15
    ["Strona www", "Strona parasolowa VFG (polowa)", "W trakcie", "czerwiec/lipiec", 15, "Podzial: 15+15"],
    # Domeny - 2
    ["Domeny", "Zakup domen dodatkowych", "W trakcie", None, 2, ""],
    # Nagrywki
    ["Nagrywki", "Wina gruzinskie (11.06)", "Zrealizowano", "11.06", None, ""],
    ["Nagrywki", "Konkurs gotowania dziennikarzy (15.06)", "Zrealizowano", "15.06", None, ""],
    ["Nagrywki", "Koncert Monika Borzym (19.06)", "Zaplanowano", "19.06", None, ""],
    ["Nagrywki", "Gala kosmetykow (25.06)", "Zaplanowano", "25.06", None, ""],
    ["Nagrywki", "LU - Hebe + kosmetyki", "Do potwierdzenia", "23.06", None, ""],
]

czerwiec_total = 0
for item in czerwiec:
    row = add_row(ws, row, item)
    if item[4]:
        czerwiec_total += item[4]

row = add_total(ws, row, "RAZEM CZERWIEC 2026:", czerwiec_total)
row += 2

# ============================================================
# LIPLEC 2026
# ============================================================
row = add_month_header(ws, row, "LIPiec 2026")

# Polowa strony www 30h = 15
# CRM - 33
# Warsztaty - 18
lipiec = [
    # Strona www - polowa
    ["Strona www", "Strona parasolowa VFG (polowa)", "W trakcie", "czerwiec/lipiec", 15, ""],
    # CRM - 33
    ["CRM", "Warsztaty CRM (przygotowanie + 2 spotkania)", "Zaplanowano", "6 i 13 lipca", 28, "przygotowanie + 2 spotkania po 5h"],
    ["CRM", "Definicja wymagan i specyfikacja", "Do realizacji", "Przed 6 lipca", 5, ""],
    # Warsztaty strategiczne - 18
    ["Warsztaty strategiczne", "Warsztat 1", "Zaplanowano", "29.06 / 6 lipca", 5, ""],
    ["Warsztaty strategiczne", "Warsztat 2", "Zaplanowano", "6 lipca / 13 lipca", 5, ""],
    ["Warsztaty strategiczne", "Spotkanie 'Kim jestesmy na rynku'", "Zaplanowano", "Lipiec-sierpien", 6, ""],
]

lipiec_total = 0
for item in lipiec:
    row = add_row(ws, row, item)
    if item[4]:
        lipiec_total += item[4]

row = add_total(ws, row, "RAZEM LIPiec 2026:", lipiec_total)
row += 2

# ============================================================
# PODSUMOWANIE
# ============================================================
podsumowanie_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
podsumowanie_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
cell = ws.cell(row=row, column=1, value="PODSUMOWANIE")
cell.font = podsumowanie_font
cell.fill = podsumowanie_fill
cell.alignment = Alignment(horizontal='center', vertical='center')
for col in range(1, 7):
    ws.cell(row=row, column=col).fill = podsumowanie_fill
    ws.cell(row=row, column=col).border = thin_border
row += 1

summary = [
    ["MAJ 2026", maj_total],
    ["CZERWIEC 2026", czerwiec_total],
    ["LIPiec 2026", lipiec_total],
]
for item in summary:
    cell_cat = ws.cell(row=row, column=1, value=item[0])
    cell_cat.font = bold_font
    cell_cat.border = thin_border
    cell_h = ws.cell(row=row, column=5, value=item[1])
    cell_h.font = bold_font
    cell_h.alignment = Alignment(horizontal='center', vertical='center')
    cell_h.border = thin_border
    for col in range(2, 7):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

row += 1
razem_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
razem_font = Font(name="Calibri", bold=True, size=16, color="006100")

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
cell = ws.cell(row=row, column=1, value="RAZEM MAJ+CZERWIEC+LIPiec:")
cell.font = razem_font
cell.fill = razem_fill
cell.border = thin_border
cell_h = ws.cell(row=row, column=5, value=maj_total + czerwiec_total + lipiec_total)
cell_h.font = razem_font
cell_h.fill = razem_fill
cell_h.alignment = Alignment(horizontal='center', vertical='center')
cell_h.border = thin_border
for col in range(1, 7):
    ws.cell(row=row, column=col).fill = razem_fill
    ws.cell(row=row, column=col).border = thin_border

ws.freeze_panes = 'A2'

output_path = r"C:\Users\Malgorzata\Desktop\GRUPA VILLA FOKSAL ANTYGRAVITY\GVF\GVF_Proces_Maj_Czerwiec_Lipiec.xlsx"
wb.save(output_path)

print("PODSUMOWANIE:")
print("="*40)
print(f"MAJ:      {maj_total}h")
print(f"CZERWIEC: {czerwiec_total}h")
print(f"LIPiec:   {lipiec_total}h")
print(f"----------------------------")
print(f"RAZEM:    {maj_total + czerwiec_total + lipiec_total}h")
print()
print("PODZIAL 'maj/czerwiec' i 'czerwiec/lipiec':")
print(f"Materialy LU 70h -> maj 35h + czerwiec 35h")
print(f"Strona www 30h -> czerwiec 15h + lipiec 15h")
