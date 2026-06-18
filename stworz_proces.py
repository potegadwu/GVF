import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Odczyt istniejącego pliku
wb_src = openpyxl.load_workbook(r'C:\Users\Malgorzata\Desktop\GRUPA VILLA FOKSAL ANTYGRAVITY\GVF\GVF_Zestawienie_Dzialan_v1 GS.xlsx')
ws_src = wb_src['Zestawienie Dzia\u0142a\u0144']

# Odczyt danych
all_data = []
for row in ws_src.iter_rows(min_row=2, max_row=ws_src.max_row, values_only=True):
    all_data.append(row)

# Tworzenie nowego pliku
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Proces Maj-Czerwiec"

# Style
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
category_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
category_font = Font(name="Calibri", bold=True, size=12)
total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
total_font = Font(name="Calibri", bold=True, size=11, color="006100")
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Nagłówki
headers = ["Kategoria", "Temat / Działanie", "Status", "Termin", "Osoba", "Decyzyjność", "Godziny (h)", "Uwagi"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Szerokości kolumn
col_widths = [35, 45, 16, 20, 20, 45, 12, 35]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

def add_category_row(ws, row_num, category_name):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
    cell = ws.cell(row=row_num, column=1, value=category_name)
    cell.font = category_font
    cell.fill = category_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, 9):
        ws.cell(row=row_num, column=col).border = thin_border
        ws.cell(row=row_num, column=col).fill = category_fill
    return row_num + 1

def add_data_row(ws, row_num, data):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 7 and isinstance(value, (int, float)):
            cell.alignment = Alignment(horizontal='center', vertical='center')
    return row_num + 1

def add_total_row(ws, row_num, label, total_hours):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal='right', vertical='center')
    for col in range(1, 9):
        ws.cell(row=row_num, column=col).border = thin_border
        ws.cell(row=row_num, column=col).fill = total_fill
    total_cell = ws.cell(row=row_num, column=7, value=total_hours)
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.alignment = Alignment(horizontal='center', vertical='center')
    return row_num + 1

row = 2

# ============================================================
# 1. CRM
# ============================================================
row = add_category_row(ws, row, "1. CRM - Prezentacja rozwiązania z rekomendacją wdrożenia")
crm_hours = 0
crm_data = [
    ["CRM", "Definicja wymagań i specyfikacja", "Do realizacji", "Przed 6 lipca", "Zespół", "Ustalenie: głębokość wpięcia w CRM, integracja z food costem, format ofert, przepływ zleceń na kuchnię", 5, "Stworzyć dokument specyfikacji przed warsztatami"],
    ["CRM", "Warsztaty do przygotowania specyfikacji i określenia możliwości i zasad wdrożenia CRM (2x spotkania)", "Zaplanowano", "6 i 13 lipca 2026", "Paweł / Marcin C / Zespół", "Specyfikacja wymagań, wybór rozwiązania (HubSpot?), integracja z food costem, format ofert", None, "Terminy w kalendarzu. Jedno spotkanie może być za mało"],
]
for item in crm_data:
    row = add_data_row(ws, row, item)
    if item[6]:
        crm_hours += item[6]
row = add_total_row(ws, row, "RAZEM CRM:", crm_hours)

# ============================================================
# 2. STRATEGIA ZMIANY NA GRUPĘ
# ============================================================
row = add_category_row(ws, row, "2. Strategia zmiany na grupę - możliwe podejścia")
strategia_hours = 0
strategia_data = [
    ["Identyfikacja grupy", "Analiza rynku, konkurencji, strategia dla VFG", "W trakcie", None, "Madeby", "Przygotowanie strategii - możliwe opcje działania i skutki wdrożenia podejść", 20, "Omawiane na spotkaniu 16.05"],
    ["Identyfikacja grupy", "Zmiana firmy na Villa Foksal Group", "W trakcie", None, "Madeby", "Wdrożenie zmian w materiałach, komunikacji", 20, "Omawiane na spotkaniu 11.05"],
    ["Warsztaty strategiczne", "Warsztat 1 - 29 czerwca", "Zaplanowano", "6 lipca 2026", "Paweł / Marcin C / Zespół", "Sprzedaż, managerowie po jednym z lokalu, Marcin C, bez kuchni", 5, "Strategia komunikacji, oferta, format, grupa docelowa, pozycjonowanie LU i VF, konkurencja"],
    ["Warsztaty strategiczne", "Warsztat 2 - 6 lipca", "Zaplanowano", "13 lipca", "Paweł / Marcin C / Zespół", "Kontynuacja prac", 5, "Cel: wypracowanie strategii podwaliny do stron, SEO, komunikacji, spotkania z zespołem"],
    ["Warsztaty strategiczne", "Spotkanie 'Kim jesteśmy na rynku' (dla zespołu)", "Zaplanowano", "Lipiec-Sierpień 2026", "Madeby", "Gotowa strategia z warsztatów", None, "Komunikacja wewnętrzna - całej strukturze pracowników"],
]
for item in strategia_data:
    row = add_data_row(ws, row, item)
    if item[6]:
        strategia_hours += item[6]
row = add_total_row(ws, row, "RAZEM STRATEGIA:", strategia_hours)

# ============================================================
# 3. WYPRACOWANIE STORY O MARCE
# ============================================================
row = add_category_row(ws, row, "3. Wypracowanie story o marce")
story_hours = 0
story_data = [
    ["Identyfikacja grupy", "Narracja grupy - 'Gościnność zapisana w historii miejsca'", "W trakcie", None, "Madeby", "Dopracowanie treści, wdrożenie na strony, mailingi, popupy, komunikacja zewnętrzna/wewnętrzna", 10, "Bazowa treść gotowa. Historia od XVIII w. przez Vauxhall, Zamoyskich, Café Bodo po dziś"],
    ["Identyfikacja grupy", "Sąsiad - Pałac na Foksal", "Zrealizowano", None, "Madeby", "Brak zagrożenia - inne miejsce, inny adres", 1, "Sprawdzone - kompletnie inne miejsce. Brak działań"],
]
for item in story_data:
    row = add_data_row(ws, row, item)
    if item[6]:
        story_hours += item[6]
row = add_total_row(ws, row, "RAZEM STORY:", story_hours)

# ============================================================
# 4. PROJEKT KONCEPCJI KREATYWNEJ GRUPY
# ============================================================
row = add_category_row(ws, row, "4. Projekt koncepcji kreatywnej Grupy Villa Foksal (logo, wizytówki, www)")
koncepcja_hours = 0
koncepcja_data = [
    ["Logo grupy", "Logo grupy - projekt finalny", "W trakcie", "Do 17.06 (wtorek)", "Madeby", "Akceptacja Klienta, przełożenie na formaty", 12, "Wybraliśmy kierunek, we wtorek finalne + formaty"],
    ["Logo grupy", "Logo na materiałach i www", "Zaplanowano", "Po akceptacji logo", "Madeby", "Gotowe logo w formatach wektorowych", 3, "Pokazanie jak logo pracuje na materiałach, www"],
    ["Wizytówki", "Wizytówki - wycena i druk", "W trakcie", "Do końca czerwca", "Gosia", "Decyzja: stanowiska na wizytówkach (Agata vs Robert/Gosia - gradacja). Pliki do drukarni", 4, "Wybrane kartony - dłuższy czas oczekiwania"],
    ["Strona www", "Strona parasolowa VillaFoksalGroup.pl", "W trakcie", "Czerwiec-Lipiec 2026", "Madeby", "Projekt statyczny strony głównej + podstrony, struktura, copy, akceptacja Klienta", 30, None],
    ["Strona www", "Statyczne projekty stron LU i VF", "W trakcie", "Czerwiec 2026", "Marcin C", "Projekt graficzny, założenia struktury, copy", None, "Do zderzenia z wiedzą i doświadczeniem zespołu"],
    ["Strona www", "Strona główna LU (projekt)", "W trakcie", "Czerwiec 2026", "Marcin C", "Akceptacja projektu", None, None],
    ["Domeny", "Zakup domen dodatkowych", "W trakcie", None, None, "Decyzja Klienta jakie domeny dodatkowe", 2, "VillaFoksalGroup.pl, GrupaVillaFoksal.pl, VillaFoksalgrupa.pl, GrupaFoksal.pl"],
]
for item in koncepcja_data:
    row = add_data_row(ws, row, item)
    if item[6]:
        koncepcja_hours += item[6]
row = add_total_row(ws, row, "RAZEM KONCEPCJA (logo, www, wizytówki):", koncepcja_hours)

# ============================================================
# 5. SESJA REKLAMOWA LU
# ============================================================
row = add_category_row(ws, row, "5. Sesja reklamowa LU - materiały")
sesja_hours = 0
sesja_data = [
    ["Materiały reklamowe LU", "Filmy reklamowe z sesji (Bartek, Gosia, Agata, Łukasz, kuchnia x2, wnętrze x2, tour LU)", "Zrealizowano (akcept)", "Maj-czerwiec 2026", "Zespół kreatywny", "Akceptacja filmów przez Klienta (wycięcie 41 piętra, dokolorowanie)", 70, "Zaakceptowane. Wycinamy 41 piętro, dokolorowujemy elementy"],
    ["Materiały reklamowe LU", "Montaż materiałów Agata, Gosia nowe wersje (podwykonawca)", "W trakcie", "Do końca czerwca", "Podwykonawca (Agata koordynuje)", "Materiał źródłowy z naszej sesji", None, "Agata montuje z podwykonawcą"],
    ["Materiały reklamowe LU", "Przygotowanie nowych formatów do kampanii Meta", "W trakcie", "Od kolejnego tygodnia po 12.06", "Madeby / Dom mediowy", "Gotowe filmy, akceptacja przed podmianką w kreacjach", 20, "Formaty reklamowe przygotowane pod Cele kampanii"],
    ["Materiały reklamowe LU", "Dźwięk do filmu z Gosią", "Do realizacji", None, "Zespół", "Dogranie dźwięku", 2, None],
    ["Nagrywki", "Wina gruzińskie (11 czerwca)", "Zrealizowano", "11.06.2026", "Zespół", "Materiał do montażu", None, "Obszerna relacja. Ogród i oranżeria. Wszyscy managerowie"],
    ["Nagrywki", "Konkurs gotowania dziennikarzy (15 czerwca)", "Zrealizowano", "15.06.2026", "Robert / Zespół", "Materiał do social media i PR", None, "Materiał poszedł w Pytanie na Śniadanie"],
    ["Nagrywki", "Koncert Monika Borzym jazz (19 czerwca)", "Zaplanowano", "19 czerwca 2026", "Zespół", "Nagranie materiału, serwowane menu", None, None],
    ["Nagrywki", "Gala kosmetyków (25 czerwca)", "Zaplanowano", "25 czerwca 2026", "Zespół", "Nagranie materiału", None, "Wjazd scena o 18, start 19:00"],
    ["Nagrywki", "LU - Hebe + kosmetyki koreańskie (23 czerwca)", "Do potwierdzenia", "23 czerwca 2026", "Zespół", "Potwierdzenie zmian", None, "Proszę dać znać jak coś się zmieni"],
    ["Nagrywki", "Wesele w LU (30 maja)", "Zrealizowano", "30.05.2026", "Zespół", "Nagranie materiału", None, None],
    ["Nagrywki", "Manager - logistyka, wyporność, elastyczność (28 maja)", "Zrealizowano", "28.05.2026", "Manager", "Nagranie wypowiedzi", None, "Dwa cateringi + impreza VF oraz jeden w LU"],
]
for item in sesja_data:
    row = add_data_row(ws, row, item)
    if item[6]:
        sesja_hours += item[6]
row = add_total_row(ws, row, "RAZEM SESJA REKLAMOWA LU + NAGRYWKI:", sesja_hours)

# ============================================================
# 6. TRZECIE URODZINY LU - KONCEPT STRATEGICZNY/KREATYWNY
# ============================================================
row = add_category_row(ws, row, "6. Trzecie urodziny LU - koncept strategiczny/kreatywny")
urodziny_hours = 0
urodziny_data = [
    ["Urodziny LU", "Prezentacja konceptu urodzinowego", "Zrealizowano", "Do 23.06 (next status)", "Paweł/ Zespół", "Akceptacja konceptu przez Klienta, ustalenie elementów głównych (drony, lasery, fundusz)", 8, "Pomysł, potrzeby, podejście omawiane na statusach 12.06 i prezentacja koncepcji kreatywnej 16.06"],
    ["Urodziny LU", "Wycena wydarzenia (drony, lasery, fundusz, gadżety)", "W trakcie", "Koniec czerwca 2026", "Gosia", "Zebranie ofert od podwykonawców, wycena gadżetów, wycena pokazu dronów, laserów", 4, "Zapytania gadżetowe wysłane, odpowiedzi z końcem kolejnego tygodnia"],
    ["Urodziny LU", "Weryfikacja terminu 7 września", "Do realizacji", "Do 23.06", "Gosia", "Sprawdzenie czy 7 września jest realny, czy trzeba przesunąć", None, "TBC"],
    ["Urodziny LU", "Zaproszenie / Save the date", "Do realizacji", "Wysyłka ~26 czerwca", "Zespół (Agata/Gosia)", "Gotowy koncept graficzny, akceptacja treści, projekt graficzny", 8, "Pierwsze wrażenie - nie spieszyć się"],
    ["Urodziny LU", "Komunikacja eventu (maile)", "Zaplanowano", "czerwiec-lipiec-sierpień 2026", "Zespół", "Gotowa lista odbiorców, treści mailingów (3 mailingi)", 10, "20 czerwiec: 1st mailing. Lipiec: przypominający. Sierpień: przypominający"],
    ["Urodziny LU", "Menu na wydarzenie", "Do realizacji", "Do 23.06 (omówienie)", "Marcin W (chef)", "Propozycja menu przez Marcina W", 4, "Marcin W ma przygotować propozycję"],
    ["Urodziny LU", "Scenariusz / Prowadzący", "Do realizacji", "Do końca czerwca", "Zespół", "Decyzja kto prowadzi (propozycja: Robert), ustalenie roli prowadzącego", 8, "Paweł przemawia, wycena prowadzącego"],
    ["Urodziny LU", "Animacje na ekranach (diody, totemy, ekrany)", "Do realizacji", "Lipiec-sierpień", "Madeby", "Scenariusz animacji - historia LU + historia grupy", 100, "Wykorzystamy diode, totem i ekrany"],
    ["Urodziny LU", "Welcome / Good Night Pack", "Do realizacji", "Lipiec-sierpień", "Gosia", "Projekt gadżetu użytkowego z logo grupy", 2, "Coś użytecznego z logo grupy"],
    ["Urodziny LU", "Rozpoznanie partnerstwa (współpraca z właścicielem budynku)", "W trakcie", "Do końca czerwca", "Paweł", "Rozmowa z właścicielem budynku o partycypacji w kosztach (np. 100k)", 1, "Paweł ma rozeznać temat"],
    ["Urodziny LU", "Partnerstwo / Sponsorzy", "Do przemyślenia", "Lipiec", "Zespół", "Znalezienie partnera, który mógłby współfinansować (20k?)", 6, "Partner pasujący, niekonkurujący"],
    ["Urodziny LU", "Budżet wydarzenia", "Do realizacji", "Do końca czerwca", "Zespół", "Policzenie kosztów, ustalenie możliwego budżetu", 6, "Potrzebny do przygotowania scenariusza"],
]
for item in urodziny_data:
    row = add_data_row(ws, row, item)
    if item[6]:
        urodziny_hours += item[6]
row = add_total_row(ws, row, "RAZEM URODZINY LU:", urodziny_hours)

# ============================================================
# 7. SOME - DANE
# ============================================================
row = add_category_row(ws, row, "7. SoMe - dane, treści, posty")
some_hours = 0
some_data = [
    ["Nowa formuła video", "Zmiana formuły materiałów video do social media", "W trakcie", "Czerwiec 2026", "Michał / Marcin C / Marcin W", "Omówienie nowej formuły, przejrzenie materiału z sesji rok temu dla VF", 10, "Spotkanie Michała z Marcinem - przegląd materiału z sesji VF"],
    ["Nowa formuła video", "Scenariusze postów na konta", "W trakcie", "Czerwiec 2026", "Michał / Marcin C / Agata", "Treści postów do akceptacji, omówienie zmian w komunikacji", 3, "Posty idą na bieżąco. Omówić z Agatą"],
    ["Nowa formuła video", "Posty z sesji zeszłorocznej w willi", "W trakcie", "Czerwiec 2026", "Zespół", "Akceptacja treści i grafik", 6, "Materiał 'roboczy', zabawny, budujący atmosferę"],
    ["Pracodawcy RP", "Śniadanie u Pracodawcy RP", "W trakcie", "18-19.05.2026", "Paweł / Zespół", "Prezentacja grupy (LU, VF, CVF, grupa)", 2, "Paweł mówi w kilku słowach o poszczególnych podmiotach"],
    ["Pracodawcy RP", "Mailing w Pracodawcy RP", "W trakcie", None, "Zespół", "Przygotowanie treści do akceptu", 2, "Będziemy w mailingu"],
    ["Pracodawcy RP", "Materiał w miesięczniku Pracodawcy RP", "W trakcie", None, "Zespół", "Przygotowanie treści", 2, "Będziemy mieć materiał w miesięczniku"],
    ["Forbes", "Współpraca z Forbes", "Oczekiwanie", None, "Zespół", "Oczekiwanie na ofertę współpracy", 1, "Jeszcze nie dał oferty"],
]
for item in some_data:
    row = add_data_row(ws, row, item)
    if item[6]:
        some_hours += item[6]
row = add_total_row(ws, row, "RAZEM SOME / PR / KOMUNIKACJA:", some_hours)

# ============================================================
# 8. MATERIAŁY VIDEO / MONTAŻE
# ============================================================
row = add_category_row(ws, row, "8. Materiały video / montaże")
video_hours = 0
video_data = [
    ["Sesja VF", "Sesja zdjęciowa - ogród i kuchnia (24 czerwca)", "Zaplanowano", "24 czerwca 2026", "Marcin C / Marcin W", "Scenariusz ujęć od Marcina C dla Marcina W, potwierdzenie terminu", None, "Marcin C przygotowuje szczegółowy scenariusz ujęć na cały dzień"],
    ["Sesja VF", "Materiał tour LU wewnętrzny", "Zrealizowano", None, "Zespół", "Oddany do dyspozycji Klienta", 8, "Dla rozmów z agencjami"],
    ["Materiał oprowadzania", "Film oprowadzania po LU w trakcie prac", "W trakcie", None, "Zespół", "Dopracowanie materiału, udostępnienie na linku", None, "Szybko okazał się wewnętrznie potrzebny"],
]
for item in video_data:
    row = add_data_row(ws, row, item)
    if item[6]:
        video_hours += item[6]
row = add_total_row(ws, row, "RAZEM MATERIAŁY VIDEO / MONTAŻE:", video_hours)

# ============================================================
# 9. KAMPANIE MEDIOWE
# ============================================================
row = add_category_row(ws, row, "9. Kampanie mediowe")
kampanie_hours = 0
kampanie_data = [
    ["Kampanie reklamowe", "Media plan 2026 - rewizja i korekty", "W trakcie", "Czerwiec 2026", "Zespół / Dom mediowy", "Przedstawienie proponowanych zmian/poszerzenia TG", 4, "100k zł netto z obsługą mediową"],
    ["Kampanie reklamowe", "Kampania Catering (cały rok)", "W trakcie", "Styczeń-Grudzień 2026", "Dom mediowy", "Miesięczna optymalizacja", None, "Ciągła kampania miesiąc w miesiąc"],
    ["Kampanie reklamowe", "Kampania Wesela (cały rok)", "W trakcie", "Styczeń-Grudzień 2026", "Dom mediowy", "Miesięczna optymalizacja", None, "Ciągła kampania miesiąc w miesiąc"],
    ["Kampanie reklamowe", "Konferencje w LU (maj-czerwiec)", "W trakcie", "Maj-Czerwiec 2026", "Dom mediowy", "Włączone z końcem kwietnia", None, "Zastąpiły 'małe konferencje'"],
    ["Kampanie reklamowe", "Garden Party VF (maj-czerwiec)", "W trakcie", "Maj-Czerwiec 2026", "Dom mediowy", "Przedłużone o dwa miesiące", None, None],
    ["Kampanie reklamowe", "Reklamy z Robertem, Bartkiem, Łukaszem, LU, kuchnia", "W trakcie", None, None, "Przygotowanie 30 do kampanii w meta", 50, None],
]
for item in kampanie_data:
    row = add_data_row(ws, row, item)
    if item[6]:
        kampanie_hours += item[6]
row = add_total_row(ws, row, "RAZEM KAMPANIE MEDIOWE:", kampanie_hours)

# ============================================================
# 10. WYPRACOWYWANIE NOWEGO PODJĘCIA DO MATERIAŁÓW W SOME
# ============================================================
row = add_category_row(ws, row, "10. Wypracowywanie nowego podejścia do materiałów prezentowanych w SoMe")
nowe_podejscie_hours = 0
nowe_podejscie_data = [
    ["Nowa formuła video", "Zmiana formuły materiałów video do social media", "W trakcie", "Czerwiec 2026", "Michał / Marcin C / Marcin W", "Omówienie nowej formuły, przejrzenie materiału z sesji rok temu dla VF", 10, "Spotkanie Michała z Marcinem - przegląd materiału z sesji VF"],
    ["Nowa formuła video", "Scenariusze postów na konta", "W trakcie", "Czerwiec 2026", "Michał / Marcin C / Agata", "Treści postów do akceptacji, omówienie zmian w komunikacji", 3, "Posty idą na bieżąco"],
    ["Nowa formuła video", "Posty z sesji zeszłorocznej w willi", "W trakcie", "Czerwiec 2026", "Zespół", "Akceptacja treści i grafik", 6, "Materiał 'roboczy', zabawny, budujący atmosferę"],
    ["Materiały reklamowe LU", "Przygotowanie nowych formatów do kampanii Meta", "W trakcie", "Od kolejnego tygodnia po 12.06", "Madeby / Dom mediowy", "Gotowe filmy, akceptacja przed podmianką w kreacjach", 20, "Formaty reklamowe przygotowane pod Cele kampanii"],
]
for item in nowe_podejscie_data:
    row = add_data_row(ws, row, item)
    if item[6]:
        nowe_podejscie_hours += item[6]
row = add_total_row(ws, row, "RAZEM NOWE PODJĘCIE SOme:", nowe_podejscie_hours)

# ============================================================
# PODSUMOWANIE OGÓLNE
# ============================================================
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
podsumowanie_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
podsumowanie_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
cell = ws.cell(row=row, column=1, value="PODSUMOWANIE")
cell.font = podsumowanie_font
cell.fill = podsumowanie_fill
cell.alignment = Alignment(horizontal='center', vertical='center')
for col in range(1, 9):
    ws.cell(row=row, column=col).fill = podsumowanie_fill
    ws.cell(row=row, column=col).border = thin_border

row += 1
podsumowanie = [
    ["1. CRM", crm_hours],
    ["2. Strategia zmiany na grupę", strategia_hours],
    ["3. Wypracowanie story o marce", story_hours],
    ["4. Projekt koncepcji kreatywnej (logo, www, wizytówki)", koncepcja_hours],
    ["5. Sesja reklamowa LU + nagrywki", sesja_hours],
    ["6. Trzecie urodziny LU", urodziny_hours],
    ["7. SoMe / PR / Komunikacja", some_hours],
    ["8. Materiały video / montaże", video_hours],
    ["9. Kampanie mediowe", kampanie_hours],
    ["10. Nowe podejście do materiałów SoMe", nowe_podejscie_hours],
]
suma_total = 0
for item in podsumowanie:
    cell_cat = ws.cell(row=row, column=1, value=item[0])
    cell_cat.font = bold_font
    cell_cat.border = thin_border
    cell_hours = ws.cell(row=row, column=7, value=item[1])
    cell_hours.font = bold_font
    cell_hours.alignment = Alignment(horizontal='center', vertical='center')
    cell_hours.border = thin_border
    for col in range(2, 9):
        ws.cell(row=row, column=col).border = thin_border
    suma_total += item[1]
    row += 1

# RAZEM
row += 1
cell_razem_label = ws.cell(row=row, column=1, value="ŁĄCZNY CZAS KOORDYNACJI:")
cell_razem_label.font = Font(name="Calibri", bold=True, size=14, color="006100")
cell_razem_label.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
cell_razem_label.border = thin_border
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
cell_razem_value = ws.cell(row=row, column=7, value=suma_total)
cell_razem_value.font = Font(name="Calibri", bold=True, size=14, color="006100")
cell_razem_value.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
cell_razem_value.alignment = Alignment(horizontal='center', vertical='center')
cell_razem_value.border = thin_border
for col in range(1, 9):
    ws.cell(row=row, column=col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws.cell(row=row, column=col).border = thin_border

# Zamrożenie
ws.freeze_panes = 'A2'

# Zapis
output_path = r"C:\Users\Malgorzata\Desktop\GRUPA VILLA FOKSAL ANTYGRAVITY\GVF\GVF_Proces_Maj_Czerwiec.xlsx"
wb.save(output_path)
print(f"Plik zapisany: {output_path}")
print(f"\nPODSUMOWANIE GODZIN:")
for item in podsumowanie:
    print(f"  {item[0]}: {item[1]}h")
print(f"\nŁĄCZNY CZAS KOORDYNACJI: {suma_total}h")
